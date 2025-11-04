"""
Excel utility functions for reading and converting Excel files to JSON format.

This module provides functions to read Excel files (.xlsx) and convert them to the
JSON format expected by collectors.
"""

import openpyxl
from typing import List, Dict, Any
from pathlib import Path


def excel_to_json(file_path: str) -> List[Dict[str, Any]]:
    """
    Convert Excel file to list of dictionaries (JSON-compatible format).

    Args:
        file_path: Path to Excel file (.xlsx)

    Returns:
        List of dictionaries where keys are column headers

    Raises:
        FileNotFoundError: If file doesn't exist
        ValueError: If file is not a valid Excel file

    Example:
        >>> products = excel_to_json("input/products.xlsx")
        >>> print(len(products))
        62
        >>> print(products[0].keys())
        dict_keys(['item_#', 'department', 'description_1', 'upc', ...])
    """
    file_path = Path(file_path)

    if not file_path.exists():
        raise FileNotFoundError(f"Excel file not found: {file_path}")

    if file_path.suffix.lower() not in ['.xlsx', '.xlsm']:
        raise ValueError(f"File must be Excel format (.xlsx or .xlsm): {file_path}")

    try:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        ws = wb.active

        # Get headers from first row
        headers = []
        for cell in ws[1]:
            headers.append(cell.value)

        # Convert rows to dictionaries
        records = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            record = {}
            for i, value in enumerate(row):
                if i < len(headers):
                    # Convert None to empty string for consistency
                    record[headers[i]] = value if value is not None else ""
            records.append(record)

        wb.close()
        return records

    except Exception as e:
        raise ValueError(f"Failed to read Excel file: {str(e)}")


def is_excel_file(file_path: str) -> bool:
    """
    Check if a file is an Excel file based on extension.

    Args:
        file_path: Path to file

    Returns:
        True if file has .xlsx or .xlsm extension

    Example:
        >>> is_excel_file("products.xlsx")
        True
        >>> is_excel_file("products.json")
        False
    """
    return Path(file_path).suffix.lower() in ['.xlsx', '.xlsm']


def load_products(file_path: str) -> List[Dict[str, Any]]:
    """
    Load products from either JSON or Excel file.

    Automatically detects file type and reads accordingly.

    Args:
        file_path: Path to JSON or Excel file

    Returns:
        List of product dictionaries

    Raises:
        FileNotFoundError: If file doesn't exist
        ValueError: If file format is not supported

    Example:
        >>> products = load_products("input/products.xlsx")
        >>> products = load_products("input/products.json")
    """
    import json

    file_path = Path(file_path)

    if not file_path.exists():
        raise FileNotFoundError(f"File not found: {file_path}")

    if is_excel_file(str(file_path)):
        return excel_to_json(str(file_path))
    elif file_path.suffix.lower() == '.json':
        with open(file_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
            if not isinstance(data, list):
                raise ValueError("JSON file must contain an array of objects")
            return data
    else:
        raise ValueError(f"Unsupported file format: {file_path.suffix}. Use .xlsx or .json")
