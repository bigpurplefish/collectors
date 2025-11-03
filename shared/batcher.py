#!/usr/bin/env python3
"""
batcher.py — Split/Merge/Group/Sort JSON, Excel ↔ JSON, and Patch Columns.

GUI + CLI
---------
- CLI subcommands: split, merge, split-by (GUI covers those + sorting)
- GUI built with ttkbootstrap (theme: flatly), Arial 11, 1200x800
- Read-only inputs for files/dirs with Browse buttons
- Validates required inputs before running
- Background thread for long-running tasks; live log panel
- Config persistence to ~/.json_batcher_config.json

Features
--------
- split:  Preserves original order; writes batch_001.json, batch_002.json, ...
- merge:  Replaces original records by ID using augmented batches; preserves order
- split-by: Split into files by (similar) values of a key
- sort:  Multi-key stable sort with ASC/DESC per key
- Excel → JSON: Flexible orient, JSONL, indent, ASCII options
- JSON → Excel: NDJSON support, optional flatten, auto column widths
- Patch Columns: Selectively overwrite/add fields from an updated JSON into the original; optional append new records

Usage (CLI)
-----------
# Split into batches of 50 (recommended default)
python batcher.py split --input records.json --outdir ./batches --size 50 --id-field "item_#" --fix-nan

# Merge augmented batches back into a single file
python batcher.py merge --original records.json --batches-dir ./augmented --output merged.json --id-field "item_#"

# NDJSON variant (if your input is line-delimited JSON objects)
python batcher.py split --input records.ndjson --outdir ./batches --size 50 --ndjson --id-field "item_#"
python batcher.py merge --original records.ndjson --batches-dir ./augmented --output merged.ndjson --ndjson --id-field "item_#"
"""
from __future__ import annotations
import argparse
import io
import os
import json
import math
import re
import sys
import tempfile
from glob import glob
from typing import Any, Dict, Iterable, List, Tuple
import unicodedata
from pathlib import Path
import csv

# ---------------------------- Core Types ----------------------------
JsonObj = Dict[str, Any]

# ---------------------------- Global Variables ----------------------------
CONFIG_FILE = os.path.expanduser("~/.json_batcher_config.json")

# ---------------------------- IO Helpers ----------------------------
def _read_text(path: str) -> str:
    with open(path, "r", encoding="utf-8") as f:
        return f.read()

def _write_atomic(path: str, data: str) -> None:
    os.makedirs(os.path.dirname(os.path.abspath(path)) or ".", exist_ok=True)
    with tempfile.NamedTemporaryFile("w", delete=False, encoding="utf-8") as tmp:
        tmp.write(data)
        tmp_path = tmp.name
    os.replace(tmp_path, path)

def _iter_df_records_in_batches(df, batch_size: int):
    """
    Yield lists of dicts (records) from a DataFrame in batches.
    """
    n = len(df)
    if n == 0:
        return
    start = 0
    while start < n:
        end = min(start + batch_size, n)
        # Using to_dict on a slice is efficient and avoids building the whole list at once
        yield df.iloc[start:end].to_dict(orient="records")
        start = end


def _stream_json_records_atomic(path: str, records_iter, ndjson: bool, indent: int | None, force_ascii: bool):
    """
    Stream records to a temp file and replace target atomically.
    - ndjson=True  → write JSON Lines (one object per line)
    - ndjson=False → write a JSON array with optional pretty indent
    """
    import tempfile, os, json

    os.makedirs(os.path.dirname(os.path.abspath(path)) or ".", exist_ok=True)

    with tempfile.NamedTemporaryFile("w", delete=False, encoding="utf-8") as tmp:
        tmp_path = tmp.name
        if ndjson:
            for batch in records_iter:
                for obj in batch:
                    tmp.write(json.dumps(obj, ensure_ascii=force_ascii))
                    tmp.write("\n")
        else:
            # JSON array streaming
            first = True
            if indent is not None:
                # pretty-printed array
                sp = " " * indent
                tmp.write("[\n")
                for batch in records_iter:
                    for obj in batch:
                        if not first:
                            tmp.write(",\n")
                        tmp.write(sp)
                        tmp.write(json.dumps(obj, ensure_ascii=force_ascii))
                        first = False
                tmp.write("\n]")
            else:
                # compact array
                tmp.write("[")
                for batch in records_iter:
                    for obj in batch:
                        if not first:
                            tmp.write(",")
                        tmp.write(json.dumps(obj, ensure_ascii=force_ascii))
                        first = False
                tmp.write("]")
    os.replace(tmp_path, path)

def _excel_upc_to_text(val):
    """
    Coerce an Excel 'upc' cell to a 12-digit string with leading zeros preserved.
    - Strips all non-digits (handles numbers like 51115052798.0, commas, spaces)
    - Returns None for blank/NaN values
    - Pads to 12 digits (zfill), which is the standard UPC-A length
    """
    # Lazy import-safe checks to avoid requiring numpy/pandas here
    try:
        # NaN check without importing numpy/pandas
        import math
        if isinstance(val, float) and math.isnan(val):
            return None
    except Exception:
        pass

    if val is None:
        return None

    s = str(val).strip()
    if not s or s.lower() in {"nan", "none"}:
        return None

    # Remove all non-digits (handles ".0", spaces, hyphens, commas)
    s = re.sub(r"\D", "", s)
    if not s:
        return None

    # Left-pad to 12 (UPC-A)
    return s.zfill(12)

# Replace bare NaN tokens with JSON null, while avoiding quoted "NaN"
_NAN_RE = re.compile(r'(?<!")\bNaN\b(?!")')
def _fix_nan_tokens(s: str) -> str:
    return _NAN_RE.sub('null', s)

def load_config():
    if os.path.exists(CONFIG_FILE):
        try:
            with open(CONFIG_FILE, "r") as f:
                return json.load(f)
        except Exception:
            pass
    return {}

def save_config(values: dict):
    try:
        with open(CONFIG_FILE, "w") as f:
            json.dump(values, f)
    except Exception:
        pass

# ---------------------------- JSON Load/Save ----------------------------
def _load_json_array(path: str, fix_nan: bool, ndjson: bool) -> List[JsonObj]:
    if ndjson:
        lines = _read_text(path).splitlines()
        out: List[JsonObj] = []
        for i, line in enumerate(lines, 1):
            if not line.strip():
                continue
            try:
                if fix_nan:
                    line = _fix_nan_tokens(line)
                out.append(json.loads(line))
            except json.JSONDecodeError as e:
                raise ValueError(f"NDJSON parse error on line {i}: {e}") from e
        return out
    else:
        text = _read_text(path)
        if fix_nan:
            text = _fix_nan_tokens(text)
        try:
            data = json.loads(text)
        except json.JSONDecodeError as e:
            raise ValueError(f"JSON parse error in {path}: {e}") from e
        if not isinstance(data, list):
            raise ValueError("Expected a JSON array as top-level structure.")
        return data

def _dump_json_array(path: str, items: List[JsonObj], ndjson: bool) -> None:
    if ndjson:
        payload = "\n".join(json.dumps(obj, ensure_ascii=False) for obj in items) + "\n"
    else:
        payload = json.dumps(items, ensure_ascii=False, indent=2)
    _write_atomic(path, payload)

# ---------------------------- Utilities ----------------------------
def _pad_width(n_items: int, size: int) -> int:
    n_batches = max(1, math.ceil(n_items / max(1, size)))
    return max(3, len(str(n_batches)))

def _validate_and_index(items: List[JsonObj], id_field: str) -> Tuple[Dict[str, int], List[str]]:
    idx: Dict[str, int] = {}
    warnings: List[str] = []
    for pos, obj in enumerate(items):
        if id_field not in obj or obj[id_field] in (None, ""):
            warnings.append(f"Record at position {pos} missing '{id_field}'.")
            continue
        key = str(obj[id_field])
        if key in idx:
            warnings.append(f"Duplicate '{id_field}'='{key}' at positions {idx[key]} and {pos}.")
        else:
            idx[key] = pos
    return idx, warnings

def _slugify(s: str, maxlen: int = 80) -> str:
    s = unicodedata.normalize("NFKD", s)
    s = "".join(ch for ch in s if not unicodedata.combining(ch))
    s = re.sub(r"[^a-zA-Z0-9._-]+", "-", s)
    s = s.strip("-._")
    return (s[:maxlen] or "_")

def _normalize_soft(s: str) -> str:
    s = unicodedata.normalize("NFKD", str(s)).casefold()
    s = "".join(ch for ch in s if not unicodedata.combining(ch))
    s = s.strip()
    s = re.sub(r"\s+", " ", s)
    s = re.sub(r"[^\w\s.-]", "", s)
    return s

def _canonicalize(value: Any, mode: str, *, regex_group: int = 1, mapping: Dict[str, str] | None = None) -> str:
    if value is None:
        return ""
    raw = str(value)
    if mode == "exact":
        return raw
    if mode == "casefold":
        return raw.casefold()
    if mode.startswith("prefix:"):
        try:
            n = int(mode.split(":", 1)[1])
        except Exception:
            n = 4
        base = _normalize_soft(raw)
        return base[:max(0, n)]
    if mode.startswith("regex:"):
        pat = mode.split(":", 1)[1]
        m = re.search(pat, raw)
        return (m.group(regex_group) if m else "")
    if mode.startswith("map:"):
        base = _normalize_soft(raw)
        if mapping:
            return mapping.get(raw, mapping.get(base, base))
        return base
    if mode == "normalize":
        return _normalize_soft(raw)
    return raw

# ---------------------------- Commands ----------------------------
def cmd_split(args: argparse.Namespace) -> None:
    items = _load_json_array(args.input, fix_nan=args.fix_nan, ndjson=args.ndjson)
    total = len(items)
    if total == 0:
        print("No records found; nothing to split.")
        return
    width = _pad_width(total, args.size)
    os.makedirs(args.outdir, exist_ok=True)
    batch_num = 0
    for i in range(0, total, args.size):
        batch = items[i:i + args.size]
        batch_num += 1
        fn = os.path.join(args.outdir, f"batch_{batch_num:0{width}d}.json")
        _dump_json_array(fn, batch, ndjson=False)
        print(f"Wrote {fn}  ({len(batch)} records)")
    if args.id_field:
        _, warns = _validate_and_index(items, args.id_field)
        if warns:
            print("\nID validation warnings:")
            for w in warns:
                print(" -", w)
    print(f"\nDone. Total records: {total}. Batches: {batch_num}. Batch size: {args.size}.")

def cmd_split_by(args: argparse.Namespace) -> None:
    items = _load_json_array(args.input, fix_nan=args.fix_nan, ndjson=args.ndjson)
    total = len(items)
    if total == 0:
        print("No records found; nothing to split.")
        return
    key = args.key
    mode = args.mode or "normalize"
    mapping = None
    if mode.startswith("map:"):
        mapping_path = mode.split(":", 1)[1]
        mapping = _load_mapping_file(mapping_path)
        if mapping is None:
            print(f"Warning: mapping file '{mapping_path}' not usable; falling back to normalize.")
            mode = "normalize"
    buckets: Dict[str, List[JsonObj]] = {}
    missing_bucket_name = args.missing_bucket or "_missing"
    for obj in items:
        val = obj.get(key)
        if val in (None, ""):
            bucket_key = missing_bucket_name
        else:
            bucket_key = _canonicalize(val, mode, mapping=mapping)
            if not bucket_key:
                bucket_key = missing_bucket_name
        buckets.setdefault(bucket_key, []).append(obj)
    os.makedirs(args.outdir, exist_ok=True)
    min_size = max(0, int(args.min_bucket)) if getattr(args, "min_bucket", 0) else 0
    kept, skipped = 0, 0
    manifest_entries: List[Tuple[str, int, str]] = []
    max_per = int(getattr(args, "max_per_file", 0) or 0)
    for bkey, group in sorted(buckets.items(), key=lambda kv: (-len(kv[1]), kv[0])):
        if min_size and len(group) < min_size:
            skipped += 1
            continue
        display_key = (bkey or missing_bucket_name)
        slug = _slugify(display_key or "missing")
        if max_per > 0:
            n_parts = (len(group) + max_per - 1) // max_per
            pad = max(3, len(str(n_parts)))
            for i in range(n_parts):
                chunk = group[i*max_per:(i+1)*max_per]
                count = len(chunk)
                filename = (f"key={slug}_part={i+1:0{pad}d}_count={count}.ndjson"
                            if args.ndjson else
                            f"key={slug}_part={i+1:0{pad}d}_count={count}.json")
                outpath = os.path.join(args.outdir, filename)
                _dump_json_array(outpath, chunk, ndjson=args.ndjson)
                print(f"Wrote {outpath}  ({count} records)")
                kept += 1
                manifest_entries.append((display_key, count, filename))
        else:
            count = len(group)
            filename = (f"key={slug}_count={count}.ndjson"
                        if args.ndjson else
                        f"key={slug}_count={count}.json")
            outpath = os.path.join(args.outdir, filename)
            _dump_json_array(outpath, group, ndjson=args.ndjson)
            print(f"Wrote {outpath}  ({count} records)")
            kept += 1
            manifest_entries.append((display_key, count, filename))
    manifest_entries.sort(key=lambda x: x[0].casefold())
    manifest_path = os.path.join(args.outdir, "key_list.csv")
    with open(manifest_path, "w", newline="", encoding="utf-8") as csvfile:
        writer = csv.writer(csvfile)
        writer.writerow(["key", "count", "file"])
        for k, c, f in manifest_entries:
            writer.writerow([k, c, f])
    print(f"Wrote {manifest_path}  ({len(manifest_entries)} files)")
    print(f"\nDone. Total records: {total}. Files written: {kept}. Buckets skipped (too small): {skipped}.")

def _read_all_batches(batches_dir: str) -> List[JsonObj]:
    paths = sorted(glob(os.path.join(batches_dir, "batch_*.json")))
    if not paths:
        raise FileNotFoundError(f"No files matching 'batch_*.json' in {batches_dir}")
    all_items: List[JsonObj] = []
    for p in paths:
        with open(p, "r", encoding="utf-8") as f:
            try:
                arr = json.load(f)
            except json.JSONDecodeError as e:
                raise ValueError(f"Failed to parse {p}: {e}") from e
        if not isinstance(arr, list):
            raise ValueError(f"{p} is not a JSON array.")
        all_items.extend(arr)
    return all_items

def cmd_merge(args: argparse.Namespace) -> None:
    original = _load_json_array(args.original, fix_nan=args.fix_nan, ndjson=args.ndjson)
    if not original:
        raise ValueError("Original file has no records.")
    id_field = args.id_field
    orig_index, orig_warns = _validate_and_index(original, id_field)
    if orig_warns:
        print("Original ID warnings:")
        for w in orig_warns:
            print(" -", w)
    augmented = _read_all_batches(args.batches_dir)
    aug_index: Dict[str, JsonObj] = {}
    dups: Dict[str, int] = {}
    missing_id_count = 0
    for obj in augmented:
        key = obj.get(id_field)
        if key in (None, ""):
            missing_id_count += 1
            continue
        key = str(key)
        if key in aug_index:
            dups[key] = dups.get(key, 1) + 1
        aug_index[key] = obj
    if dups:
        print("\nWarning: duplicate IDs in augmented batches:")
        for k, n in dups.items():
            print(f" - {k}: {n} occurrences (keeping the last one)")
    if missing_id_count:
        print(f"\nWarning: {missing_id_count} augmented records missing '{id_field}' were ignored.")
    merged: List[JsonObj] = []
    replaced = 0
    not_found = 0
    for obj in original:
        key = obj.get(id_field)
        key = str(key) if key not in (None, "") else None
        if key and key in aug_index:
            merged.append(aug_index[key])
            replaced += 1
        else:
            merged.append(obj)
            if key and key not in aug_index:
                not_found += 1
    extras = [k for k in aug_index.keys() if k not in orig_index]
    if extras:
        print(f"\nNote: {len(extras)} augmented record(s) did not exist in original and were not added.")
    _dump_json_array(args.output, merged, ndjson=args.ndjson)
    print(f"\nMerged → {args.output}")
    print(f"Original records   : {len(original)}")
    print(f"Augmented records  : {len(augmented)}")
    print(f"Replaced (by ID)   : {replaced}")
    print(f"Not found in aug   : {not_found}")
    print(f"Extras in aug      : {len(extras)}")

# ---------------------------- Sorting Helpers & Command ----------------------------
def _get_by_path(obj: JsonObj, path: str):
    cur = obj
    for part in path.split("."):
        if not isinstance(cur, dict) or part not in cur:
            return None
        cur = cur[part]
    return cur

def _set_by_path(obj: JsonObj, path: str, value: Any) -> None:
    parts = path.split(".")
    cur = obj
    for p in parts[:-1]:
        if not isinstance(cur.get(p), dict):
            cur[p] = {}
        cur = cur[p]
    cur[parts[-1]] = value

def _multi_key_sort(items: List[JsonObj], specs: List[Tuple[str, bool]]) -> List[JsonObj]:
    data = list(items)
    for key_path, desc in reversed(specs):
        data.sort(key=lambda o: (_get_by_path(o, key_path) if _get_by_path(o, key_path) is not None else ""),
                  reverse=bool(desc))
    return data

def cmd_sort(args: argparse.Namespace) -> None:
    items = _load_json_array(args.input, fix_nan=args.fix_nan, ndjson=args.ndjson)
    if not items:
        print("No records found; nothing to sort.")
        _dump_json_array(args.output, items, ndjson=args.ndjson)
        return
    if not getattr(args, "keys", None):
        raise ValueError("No sort keys specified.")
    if len(args.keys) != len(args.desc):
        raise ValueError("keys and desc must be same length.")
    specs = list(zip(args.keys, [bool(x) for x in args.desc]))
    sorted_items = _multi_key_sort(items, specs)
    _dump_json_array(args.output, sorted_items, ndjson=args.ndjson)
    print(f"Sorted → {args.output}")
    print(f"Total records: {len(items)}")
    for i, (k, d) in enumerate(specs, 1):
        print(f"Key {i}: {k}  order={'DESC' if d else 'ASC'}")

# ---------------------------- GUI ----------------------------
def launch_gui():
    """
    Main Tkinter GUI with tabs:
      - Split
      - Split by Key
      - Merge
      - Sort
      - Excel → JSON
      - JSON → Excel
      - Patch Columns  (NEW)
    """
    import json
    import os
    import tkinter as tk
    from tkinter import filedialog, messagebox
    from tkinter.scrolledtext import ScrolledText
    import ttkbootstrap as tb

    # --- threading + stdout capture ---
    import threading, queue, contextlib, io, sys
    _log_queue = queue.Queue()
    _active_log_widget = None

    def _set_active_log(widget):
        nonlocal _active_log_widget
        _active_log_widget = widget

    def _drain_log_queue():
        try:
            while True:
                line = _log_queue.get_nowait()
                if _active_log_widget is not None:
                    _active_log_widget.configure(state="normal")
                    _active_log_widget.insert("end", line)
                    _active_log_widget.see("end")
                    _active_log_widget.configure(state="disabled")
        except queue.Empty:
            pass
        app.after(75, _drain_log_queue)

    class _LogWriter(io.TextIOBase):
        def write(self, s): 
            if s: _log_queue.put(s)
            return len(s)
        def flush(self): pass

    @contextlib.contextmanager
    def _redirect_output_to_log():
        old_out, old_err = sys.stdout, sys.stderr
        lw = _LogWriter()
        sys.stdout, sys.stderr = lw, lw
        try:
            yield
        finally:
            sys.stdout, sys.stderr = old_out, old_err

    def _run_in_thread(target, *, on_done=None):
        def _wrapper():
            try:
                with _redirect_output_to_log():
                    target()
            except Exception as e:
                _log_queue.put(f"\n[error] {e}\n")
            finally:
                if on_done:
                    app.after(0, on_done)
        threading.Thread(target=_wrapper, daemon=True).start()

    # Load persisted configuration
    cfg = load_config() or {}

    # ---------------- Helper functions (shared) ----------------
    def _read_file(path):
        with open(path, "r", encoding="utf-8") as f:
            return f.read()

    def _fix_nan_tokens_local(text):
        import re
        return re.sub(r'(?<!")\bNaN\b(?!")', 'null', text)

    def _extract_keys_for_dropdown(path):
        try:
            text = _read_file(path)
        except Exception:
            return []
        text_fixed = _fix_nan_tokens_local(text)
        keys = set()
        try:
            data = json.loads(text_fixed)
            if isinstance(data, list) and data:
                sample = data if len(data) <= 200 else data[:200]
                for obj in sample:
                    if isinstance(obj, dict):
                        keys.update(obj.keys())
                return sorted(keys)
        except Exception:
            pass
        for i, line in enumerate(text_fixed.splitlines()):
            if i >= 400:
                break
            try:
                obj = json.loads(line)
                if isinstance(obj, dict):
                    keys.update(obj.keys())
            except Exception:
                continue
        return sorted(keys)

    def _populate_id_combobox(combo, var, keys, preferred=None):
        combo["values"] = keys
        if not keys:
            var.set("")
            return
        if preferred and preferred in keys:
            var.set(preferred)
        else:
            var.set(keys[0])

    # save core (split/merge) config
    def _save_config_now():
        cfg.update({
            "input_file": input_file_var.get(),
            "output_dir": output_dir_var.get(),
            "batch_size": batch_size_var.get(),
            "id_field_split": id_field_var_split.get(),
            "ndjson_split": bool(ndjson_var_split.get()),
            "fix_nan_split": bool(fix_nan_var_split.get()),
            "original_file": original_file_var.get(),
            "batches_dir": batches_dir_var.get(),
            "output_file": output_file_var.get(),
            "id_field_merge": id_field_var_merge.get(),
            "ndjson_merge": bool(ndjson_var_merge.get()),
            "fix_nan_merge": bool(fix_nan_var_merge.get()),
        })
        save_config(cfg)

    # Browsers
    def browse_file(var, *, id_combo=None, id_var=None, cfg_key_for_id=None):
        path = filedialog.askopenfilename()
        if not path: return
        var.set(path)
        if id_combo is not None and id_var is not None:
            try:
                keys = _extract_keys_for_dropdown(path)
                preferred = cfg.get(cfg_key_for_id) if cfg_key_for_id else None
                _populate_id_combobox(id_combo, id_var, keys, preferred)
            except Exception as e:
                messagebox.showerror("Error", f"Failed to inspect file for ID keys.\n\n{e}")
        _save_config_now()

    def browse_dir(var):
        path = filedialog.askdirectory()
        if path:
            var.set(path); _save_config_now()

    def browse_save(var):
        path = filedialog.asksaveasfilename(defaultextension=".json",
                                            filetypes=[("JSON", "*.json"), ("All Files", "*.*")])
        if path:
            var.set(path); _save_config_now()

    def browse_json_file(var):
        path = filedialog.askopenfilename(
            filetypes=[("JSON files", "*.json"), ("NDJSON files", "*.ndjson"), ("All files", "*.*")]
        )
        if not path: return
        var.set(path)
        _save_sort_cfg()

    def browse_json_save(var, use_ndjson_var=None):
        default_ext = ".ndjson" if (use_ndjson_var and use_ndjson_var.get()) else ".json"
        path = filedialog.asksaveasfilename(defaultextension=default_ext,
                                            filetypes=[("JSON", "*.json"), ("NDJSON", "*.ndjson"), ("All files", "*.*")])
        if not path: return
        var.set(path)
        _save_sort_cfg()

    def browse_json_or_jsonl_save(var, *, use_jsonl_var):
        default_ext = ".jsonl" if use_jsonl_var.get() else ".json"
        path = filedialog.asksaveasfilename(
            defaultextension=default_ext,
            filetypes=[("JSON", "*.json"), ("JSON Lines", "*.jsonl"), ("All files", "*.*")]
        )
        if path:
            var.set(path)
            _save_excel_cfg()

    # JSON → Excel helpers
    def _j2x_browse_json_input(var):
        path = filedialog.askopenfilename(filetypes=[("JSON/NDJSON files", "*.json *.ndjson"), ("All files", "*.*")])
        if path:
            var.set(path); _save_j2x_cfg()

    def _j2x_browse_excel_save(var):
        path = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                            filetypes=[("Excel Workbook", "*.xlsx"), ("All files", "*.*")])
        if path:
            var.set(path); _save_j2x_cfg()

    # ------------- Logging helper (default target set later) -------------
    def append_log(text):
        split_log.configure(state="normal")
        split_log.insert("end", text + "\n")
        split_log.see("end")
        split_log.configure(state="disabled")

    # ---------------- Command runners (split, split-by, merge, sort) ----------------
    def run_split():
        if not input_file_var.get():
            messagebox.showwarning("Missing Input", "Please choose an input file."); return
        if not output_dir_var.get():
            messagebox.showwarning("Missing Output Dir", "Please choose an output directory."); return
        try:
            size = int(batch_size_var.get())
            if size <= 0: raise ValueError
        except Exception:
            messagebox.showwarning("Invalid Size", "Batch size must be a positive integer."); return
        _save_config_now()
        args = argparse.Namespace(
            input=input_file_var.get(),
            outdir=output_dir_var.get(),
            size=size,
            id_field=id_field_var_split.get() or "item_#",
            ndjson=bool(ndjson_var_split.get()),
            fix_nan=bool(fix_nan_var_split.get()),
            func=cmd_split,
        )
        def _work():
            print("[split] Starting…\n"); cmd_split(args); print("\n[split] Completed successfully.\n")
        _run_in_thread(_work)

    def run_split_by():
        if not splitby_input_file_var.get():
            messagebox.showwarning("Missing Input", "Please choose an input file."); return
        if not splitby_outdir_var.get():
            messagebox.showwarning("Missing Output Dir", "Please choose an output directory."); return
        if not splitby_key_var.get().strip():
            messagebox.showwarning("Missing Key", "Please choose a key to group by."); return
        mode_choice = splitby_mode_var.get()
        if mode_choice == "exact": mode_str = "exact"
        elif mode_choice == "casefold": mode_str = "casefold"
        elif mode_choice == "normalize": mode_str = "normalize"
        elif mode_choice == "prefix":
            n = splitby_prefix_n_var.get().strip()
            if not n.isdigit() or int(n) < 0:
                messagebox.showwarning("Invalid Prefix Length", "Prefix length must be a non-negative integer."); return
            mode_str = f"prefix:{int(n)}"
        elif mode_choice == "regex":
            pat = splitby_regex_var.get().strip()
            if not pat:
                messagebox.showwarning("Missing Regex", "Please enter a regex pattern."); return
            mode_str = f"regex:{pat}"
        else:
            mp = splitby_mapping_file_var.get().strip()
            if not mp:
                messagebox.showwarning("Missing Mapping File", "Please choose a JSON mapping file."); return
            mode_str = f"map:{mp}"
        try:
            min_bucket = int(splitby_min_bucket_var.get())
            if min_bucket < 0: raise ValueError
        except Exception:
            messagebox.showwarning("Invalid Minimum", "Minimum bucket size must be a non-negative integer."); return
        raw_max = splitby_max_per_file_var.get().strip()
        if raw_max == "": max_per_file = 0
        else:
            if not raw_max.isdigit() or int(raw_max) <= 0:
                messagebox.showwarning("Invalid Max Per File", "Maximum records per file must be positive (or blank)."); return
            max_per_file = int(raw_max)
        _save_splitby_cfg()
        args = argparse.Namespace(
            input=splitby_input_file_var.get(),
            outdir=splitby_outdir_var.get(),
            key=splitby_key_var.get().strip(),
            mode=mode_str,
            min_bucket=min_bucket,
            missing_bucket=splitby_missing_bucket_var.get().strip() or "_missing",
            ndjson=bool(ndjson_var_splitby.get()),
            fix_nan=bool(fix_nan_var_splitby.get()),
            max_per_file=max_per_file,
        )
        _set_active_log(splitby_log)
        def _work():
            print("[split-by] Starting…\n"); cmd_split_by(args); print("\n[split-by] Completed successfully.\n")
        _run_in_thread(_work)

    def run_merge():
        if not original_file_var.get():
            messagebox.showwarning("Missing Original", "Please choose the original file."); return
        if not batches_dir_var.get():
            messagebox.showwarning("Missing Batches Dir", "Please choose the augmented batches directory."); return
        if not output_file_var.get():
            messagebox.showwarning("Missing Output File", "Please choose the output merged file path."); return
        _save_config_now()
        args = argparse.Namespace(
            original=original_file_var.get(),
            batches_dir=batches_dir_var.get(),
            output=output_file_var.get(),
            id_field=id_field_var_merge.get() or "item_#",
            ndjson=bool(ndjson_var_merge.get()),
            fix_nan=bool(fix_nan_var_merge.get()),
            func=cmd_merge,
        )
        def _work():
            print("[merge] Starting…\n"); cmd_merge(args); print("\n[merge] Completed successfully.\n")
        _run_in_thread(_work)

    def run_sort():
        if not sort_input_file_var.get():
            messagebox.showwarning("Missing Input", "Please choose an input file."); return
        if not sort_output_file_var.get():
            messagebox.showwarning("Missing Output", "Please choose the output file path."); return
        keys, descs = [], []
        for row in sort_key_rows:
            k = row["key_var"].get().strip()
            if k: 
                keys.append(k); descs.append(bool(row["desc_var"].get()))
        if not keys:
            messagebox.showwarning("No Keys", "Please add at least one sort key."); return
        args = argparse.Namespace(
            input=sort_input_file_var.get(),
            output=sort_output_file_var.get(),
            ndjson=bool(ndjson_var_sort.get()),
            fix_nan=bool(fix_nan_var_sort.get()),
            keys=keys,
            desc=descs,
        )
        _set_active_log(sort_log)
        def _work():
            print("[sort] Starting…\n"); cmd_sort(args); print("\n[sort] Completed successfully.\n")
        _run_in_thread(_work)

    # ---------------- UI setup ----------------
    app = tb.Window(themename="flatly")
    app.title("JSON Management Tool")
    app.geometry("1200x800")
    app.option_add("*Font", ("Arial", 11))

    nb = tb.Notebook(app)
    nb.pack(fill="both", expand=True, padx=10, pady=10)

    # ---------------- Split Tab ----------------
    split_tab = tb.Frame(nb)
    nb.add(split_tab, text="Split")
    split_tab.columnconfigure(1, weight=1)
    split_tab.rowconfigure(7, weight=1)

    input_file_var = tk.StringVar(value=cfg.get("input_file", ""))
    output_dir_var = tk.StringVar(value=cfg.get("output_dir", ""))
    id_field_var_split = tk.StringVar(value=cfg.get("id_field_split", ""))
    batch_size_var = tk.StringVar(value=cfg.get("batch_size", "50"))
    ndjson_var_split = tk.BooleanVar(value=bool(cfg.get("ndjson_split", False)))
    fix_nan_var_split = tk.BooleanVar(value=bool(cfg.get("fix_nan_split", False)))

    tb.Label(split_tab, text="Input JSON File:").grid(row=0, column=0, sticky="w")
    tb.Entry(split_tab, textvariable=input_file_var, state="readonly").grid(row=0, column=1, sticky="we", padx=5)
    id_field_combo_split = tb.Combobox(split_tab, textvariable=id_field_var_split, state="readonly")
    tb.Button(split_tab, text="Browse",
              command=lambda: browse_file(input_file_var, id_combo=id_field_combo_split,
                                          id_var=id_field_var_split, cfg_key_for_id="id_field_split")).grid(row=0, column=2)

    tb.Label(split_tab, text="Output Directory:").grid(row=1, column=0, sticky="w")
    tb.Entry(split_tab, textvariable=output_dir_var, state="readonly").grid(row=1, column=1, sticky="we", padx=5)
    tb.Button(split_tab, text="Browse", command=lambda: browse_dir(output_dir_var)).grid(row=1, column=2)

    tb.Label(split_tab, text="ID Field:").grid(row=2, column=0, sticky="w")
    id_field_combo_split.grid(row=2, column=1, sticky="w")

    tb.Label(split_tab, text="Batch Size:").grid(row=3, column=0, sticky="w")
    tb.Entry(split_tab, textvariable=batch_size_var, width=10).grid(row=3, column=1, sticky="w")

    tb.Checkbutton(split_tab, text="NDJSON (one object per line)", variable=ndjson_var_split)\
        .grid(row=4, column=0, sticky="w", padx=5, pady=(5, 0))
    tb.Checkbutton(split_tab, text="Fix bare NaN → null", variable=fix_nan_var_split)\
        .grid(row=5, column=0, sticky="w", padx=5, pady=(0, 5))
    tb.Button(split_tab, text="Run Split", bootstyle="primary", command=run_split)\
        .grid(row=6, column=0, columnspan=3, pady=10, sticky="w")

    log_frame = tb.Labelframe(split_tab, text="Log")
    log_frame.grid(row=7, column=0, columnspan=3, sticky="nsew", padx=5, pady=5)
    split_log = ScrolledText(log_frame, height=10, state="disabled")
    split_log.pack(fill="both", expand=True, padx=5, pady=5)
    _set_active_log(split_log)

    if input_file_var.get() and os.path.exists(input_file_var.get()):
        try:
            keys = _extract_keys_for_dropdown(input_file_var.get())
            _populate_id_combobox(id_field_combo_split, id_field_var_split, keys, cfg.get("id_field_split"))
        except Exception:
            pass
    id_field_combo_split.bind("<<ComboboxSelected>>", lambda e: _save_config_now())

    # ---------------- Split by Key Tab ----------------
    splitby_tab = tb.Frame(nb)
    nb.add(splitby_tab, text="Split by Key")
    splitby_tab.columnconfigure(1, weight=1)
    splitby_tab.rowconfigure(11, weight=1)

    splitby_input_file_var     = tk.StringVar(value=cfg.get("splitby_input_file", ""))
    splitby_outdir_var         = tk.StringVar(value=cfg.get("splitby_outdir", ""))
    splitby_key_var            = tk.StringVar(value=cfg.get("splitby_key", ""))
    splitby_mode_var           = tk.StringVar(value=cfg.get("splitby_mode_choice", "normalize"))
    splitby_prefix_n_var       = tk.StringVar(value=str(cfg.get("splitby_prefix_n", "3")))
    splitby_regex_var          = tk.StringVar(value=cfg.get("splitby_regex", r""))
    splitby_mapping_file_var   = tk.StringVar(value=cfg.get("splitby_mapping_file", ""))
    splitby_min_bucket_var     = tk.StringVar(value=str(cfg.get("splitby_min_bucket", "0")))
    splitby_max_per_file_var   = tk.StringVar(value=cfg.get("splitby_max_per_file", ""))
    splitby_missing_bucket_var = tk.StringVar(value=cfg.get("splitby_missing_bucket", "_missing"))
    ndjson_var_splitby         = tk.BooleanVar(value=bool(cfg.get("splitby_ndjson", False)))
    fix_nan_var_splitby        = tk.BooleanVar(value=bool(cfg.get("splitby_fix_nan", False)))

    def _save_splitby_cfg():
        cfg.update({
            "splitby_input_file": splitby_input_file_var.get(),
            "splitby_outdir": splitby_outdir_var.get(),
            "splitby_key": splitby_key_var.get(),
            "splitby_mode_choice": splitby_mode_var.get(),
            "splitby_prefix_n": splitby_prefix_n_var.get(),
            "splitby_regex": splitby_regex_var.get(),
            "splitby_mapping_file": splitby_mapping_file_var.get(),
            "splitby_min_bucket": splitby_min_bucket_var.get(),
            "splitby_max_per_file": splitby_max_per_file_var.get(),
            "splitby_missing_bucket": splitby_missing_bucket_var.get(),
            "splitby_ndjson": bool(ndjson_var_splitby.get()),
            "splitby_fix_nan": bool(fix_nan_var_splitby.get()),
        })
        save_config(cfg)

    tb.Label(splitby_tab, text="Input JSON File:").grid(row=0, column=0, sticky="w")
    tb.Entry(splitby_tab, textvariable=splitby_input_file_var, state="readonly").grid(row=0, column=1, sticky="we", padx=5)
    tb.Button(splitby_tab, text="Browse", command=lambda: (browse_json_file(splitby_input_file_var), _save_splitby_cfg())).grid(row=0, column=2)

    tb.Label(splitby_tab, text="Output Directory:").grid(row=1, column=0, sticky="w")
    tb.Entry(splitby_tab, textvariable=splitby_outdir_var, state="readonly").grid(row=1, column=1, sticky="we", padx=5)
    tb.Button(splitby_tab, text="Browse", command=lambda: (browse_dir(splitby_outdir_var), _save_splitby_cfg())).grid(row=1, column=2)

    tb.Label(splitby_tab, text="Group By Key:").grid(row=2, column=0, sticky="w")
    splitby_key_combo = tb.Combobox(splitby_tab, textvariable=splitby_key_var, state="normal")
    splitby_key_combo.grid(row=2, column=1, sticky="we", padx=5)
    tb.Button(splitby_tab, text="Refresh Keys", command=lambda: (
        splitby_key_combo.configure(values=_extract_keys_for_dropdown(splitby_input_file_var.get()) if (splitby_input_file_var.get() and os.path.exists(splitby_input_file_var.get())) else []),
        _save_splitby_cfg()
    )).grid(row=2, column=2)

    tb.Label(splitby_tab, text="Similarity Mode:").grid(row=3, column=0, sticky="w")
    mode_combo = tb.Combobox(splitby_tab, textvariable=splitby_mode_var, state="readonly",
                             values=["normalize", "exact", "casefold", "prefix", "regex", "map"], width=14)
    mode_combo.grid(row=3, column=1, sticky="w", padx=5)

    mode_opts_frame = tb.Frame(splitby_tab)
    mode_opts_frame.grid(row=4, column=0, columnspan=3, sticky="we", padx=5)
    mode_opts_frame.columnconfigure(1, weight=1)

    prefix_frame = tb.Frame(mode_opts_frame)
    tb.Label(prefix_frame, text="Prefix length:").grid(row=0, column=0, sticky="w", padx=(0,6))
    tb.Entry(prefix_frame, textvariable=splitby_prefix_n_var, width=8).grid(row=0, column=1, sticky="w")

    regex_frame = tb.Frame(mode_opts_frame)
    tb.Label(regex_frame, text="Regex pattern:").grid(row=0, column=0, sticky="w", padx=(0,6))
    tb.Entry(regex_frame, textvariable=splitby_regex_var).grid(row=0, column=1, sticky="we")

    map_frame = tb.Frame(mode_opts_frame)
    tb.Label(map_frame, text="Mapping JSON:").grid(row=0, column=0, sticky="w", padx=(0,6))
    tb.Entry(map_frame, textvariable=splitby_mapping_file_var, state="readonly").grid(row=0, column=1, sticky="we")
    tb.Button(map_frame, text="Browse", command=lambda: (browse_json_file(splitby_mapping_file_var), _save_splitby_cfg())).grid(row=0, column=2, padx=(6,0))

    def _update_mode_controls(*_):
        for fr in (prefix_frame, regex_frame, map_frame):
            fr.grid_forget()
        m = splitby_mode_var.get()
        if m == "prefix": prefix_frame.grid(row=0, column=0, sticky="w", pady=(4,0))
        elif m == "regex": regex_frame.grid(row=0, column=0, sticky="we", pady=(4,0))
        elif m == "map":   map_frame.grid(row=0, column=0, sticky="we", pady=(4,0))
        _save_splitby_cfg()
    mode_combo.bind("<<ComboboxSelected>>", _update_mode_controls)
    _update_mode_controls()

    tb.Label(splitby_tab, text="Minimum bucket size:").grid(row=5, column=0, sticky="w")
    tb.Entry(splitby_tab, textvariable=splitby_min_bucket_var, width=10).grid(row=5, column=1, sticky="w", padx=5)

    tb.Label(splitby_tab, text="Maximum records per file (optional):").grid(row=6, column=0, sticky="w")
    tb.Entry(splitby_tab, textvariable=splitby_max_per_file_var, width=10).grid(row=6, column=1, sticky="w", padx=5)

    tb.Label(splitby_tab, text="Missing/Empty bucket name:").grid(row=7, column=0, sticky="w")
    tb.Entry(splitby_tab, textvariable=splitby_missing_bucket_var, width=20).grid(row=7, column=1, sticky="w", padx=5)

    tb.Checkbutton(splitby_tab, text="NDJSON (one object per line)", variable=ndjson_var_splitby, command=_save_splitby_cfg)\
        .grid(row=8, column=0, sticky="w", padx=5, pady=(8,0))
    tb.Checkbutton(splitby_tab, text="Fix bare NaN → null", variable=fix_nan_var_splitby, command=_save_splitby_cfg)\
        .grid(row=9, column=0, sticky="w", padx=5, pady=(0,8))

    tb.Button(splitby_tab, text="Run Split by Key", bootstyle="warning", command=run_split_by)\
        .grid(row=10, column=0, columnspan=3, pady=10, sticky="w")

    splitby_log_frame = tb.Labelframe(splitby_tab, text="Split-by Log")
    splitby_log_frame.grid(row=11, column=0, columnspan=3, sticky="nsew", padx=5, pady=5)
    splitby_log = ScrolledText(splitby_log_frame, height=10, state="disabled")
    splitby_log.pack(fill="both", expand=True, padx=5, pady=5)

    def _splitby_refresh_keys(*_):
        path = splitby_input_file_var.get()
        vals = []
        if path and os.path.exists(path):
            try: vals = _extract_keys_for_dropdown(path)
            except Exception: pass
        splitby_key_combo.configure(values=vals)
        _save_splitby_cfg()
    splitby_input_file_var.trace_add("write", _splitby_refresh_keys)
    splitby_outdir_var.trace_add("write", lambda *_: _save_splitby_cfg())
    splitby_key_var.trace_add("write", lambda *_: _save_splitby_cfg())

    # ---------------- Merge Tab ----------------
    merge_tab = tb.Frame(nb)
    nb.add(merge_tab, text="Merge")
    merge_tab.columnconfigure(1, weight=1)

    original_file_var = tk.StringVar(value=cfg.get("original_file", ""))
    batches_dir_var = tk.StringVar(value=cfg.get("batches_dir", ""))
    output_file_var = tk.StringVar(value=cfg.get("output_file", ""))
    id_field_var_merge = tk.StringVar(value=cfg.get("id_field_merge", ""))
    ndjson_var_merge = tk.BooleanVar(value=bool(cfg.get("ndjson_merge", False)))
    fix_nan_var_merge = tk.BooleanVar(value=bool(cfg.get("fix_nan_merge", False)))

    tb.Label(merge_tab, text="Original JSON File:").grid(row=0, column=0, sticky="w")
    tb.Entry(merge_tab, textvariable=original_file_var, state="readonly").grid(row=0, column=1, sticky="we", padx=5)
    id_field_combo_merge = tb.Combobox(merge_tab, textvariable=id_field_var_merge, state="readonly")
    tb.Button(merge_tab, text="Browse",
              command=lambda: browse_file(original_file_var, id_combo=id_field_combo_merge,
                                          id_var=id_field_var_merge, cfg_key_for_id="id_field_merge")).grid(row=0, column=2)

    tb.Label(merge_tab, text="Batches Directory:").grid(row=1, column=0, sticky="w")
    tb.Entry(merge_tab, textvariable=batches_dir_var, state="readonly").grid(row=1, column=1, sticky="we", padx=5)
    tb.Button(merge_tab, text="Browse", command=lambda: browse_dir(batches_dir_var)).grid(row=1, column=2)

    tb.Label(merge_tab, text="Output File:").grid(row=2, column=0, sticky="w")
    tb.Entry(merge_tab, textvariable=output_file_var, state="readonly").grid(row=2, column=1, sticky="we", padx=5)
    tb.Button(merge_tab, text="Browse", command=lambda: browse_save(output_file_var)).grid(row=2, column=2)

    tb.Label(merge_tab, text="ID Field:").grid(row=3, column=0, sticky="w")
    id_field_combo_merge.grid(row=3, column=1, sticky="w")

    tb.Checkbutton(merge_tab, text="NDJSON", variable=ndjson_var_merge)\
        .grid(row=4, column=0, sticky="w", padx=5, pady=(5, 0))
    tb.Checkbutton(merge_tab, text="Fix bare NaN → null", variable=fix_nan_var_merge)\
        .grid(row=5, column=0, sticky="w", padx=5, pady=(0, 5))

    tb.Button(merge_tab, text="Run Merge", bootstyle="success", command=run_merge)\
        .grid(row=6, column=0, columnspan=3, pady=10, sticky="w")

    if original_file_var.get() and os.path.exists(original_file_var.get()):
        try:
            keys = _extract_keys_for_dropdown(original_file_var.get())
            _populate_id_combobox(id_field_combo_merge, id_field_var_merge, keys, cfg.get("id_field_merge"))
        except Exception:
            pass
    id_field_combo_merge.bind("<<ComboboxSelected>>", lambda e: _save_config_now())

    # ---------------- Sort Tab ----------------
    sort_tab = tb.Frame(nb)
    nb.add(sort_tab, text="Sort")
    sort_tab.columnconfigure(1, weight=1)
    sort_tab.rowconfigure(8, weight=1)

    sort_input_file_var = tk.StringVar(value=cfg.get("sort_input_file", ""))
    sort_output_file_var = tk.StringVar(value=cfg.get("sort_output_file", ""))
    ndjson_var_sort = tk.BooleanVar(value=bool(cfg.get("ndjson_sort", False)))
    fix_nan_var_sort = tk.BooleanVar(value=bool(cfg.get("fix_nan_sort", False)))

    def _save_sort_cfg():
        specs = []
        for rec in sort_key_rows:
            k = (rec["key_var"].get() or "").strip()
            d = bool(rec["desc_var"].get())
            if k: specs.append({"key": k, "desc": d})
        cfg.update({
            "sort_input_file": sort_input_file_var.get(),
            "sort_output_file": sort_output_file_var.get(),
            "ndjson_sort": bool(ndjson_var_sort.get()),
            "fix_nan_sort": bool(fix_nan_var_sort.get()),
            "sort_specs": specs,
        })
        save_config(cfg)

    tb.Label(sort_tab, text="Input JSON File:").grid(row=0, column=0, sticky="w")
    tb.Entry(sort_tab, textvariable=sort_input_file_var, state="readonly").grid(row=0, column=1, sticky="we", padx=5)
    tb.Button(sort_tab, text="Browse", command=lambda: browse_json_file(sort_input_file_var)).grid(row=0, column=2)

    tb.Label(sort_tab, text="Output File:").grid(row=1, column=0, sticky="w")
    tb.Entry(sort_tab, textvariable=sort_output_file_var, state="readonly").grid(row=1, column=1, sticky="we", padx=5)
    tb.Button(sort_tab, text="Browse", command=lambda: browse_json_save(sort_output_file_var, use_ndjson_var=ndjson_var_sort)).grid(row=1, column=2)

    tb.Checkbutton(sort_tab, text="NDJSON (one object per line)", variable=ndjson_var_sort, command=_save_sort_cfg)\
        .grid(row=2, column=0, sticky="w", padx=5, pady=(5, 0))
    tb.Checkbutton(sort_tab, text="Fix bare NaN → null", variable=fix_nan_var_sort, command=_save_sort_cfg)\
        .grid(row=3, column=0, sticky="w", padx=5, pady=(0, 5))

    tb.Label(sort_tab, text="Sort Keys (top = highest priority):").grid(row=4, column=0, sticky="w", pady=(10, 0))

    keys_frame = tb.Frame(sort_tab)
    keys_frame.grid(row=5, column=0, columnspan=3, sticky="we", padx=5)
    keys_frame.columnconfigure(1, weight=1)

    sort_key_rows: List[Dict[str, Any]] = []
    def add_sort_key_row(prefill_key: str = "", prefill_desc: bool = False):
        row_index = len(sort_key_rows)
        frame = tb.Frame(keys_frame)
        frame.grid(row=row_index, column=0, sticky="we", pady=2)
        frame.columnconfigure(1, weight=1)
        key_var = tk.StringVar(value=prefill_key)
        desc_var = tk.BooleanVar(value=prefill_desc)
        tb.Label(frame, text=f"Key {row_index + 1}:").grid(row=0, column=0, sticky="w", padx=(0, 6))
        key_combo = tb.Combobox(frame, textvariable=key_var, state="normal")
        key_combo.grid(row=0, column=1, sticky="we")
        order_var = tk.StringVar(value="Descending" if prefill_desc else "Ascending")
        def sync_order_to_bool(*_):
            desc_var.set(order_var.get() == "Descending"); _save_sort_cfg()
        order_var.trace_add("write", sync_order_to_bool)
        order_combo = tb.Combobox(frame, textvariable=order_var, state="readonly", values=["Ascending", "Descending"], width=12)
        order_combo.grid(row=0, column=2, padx=6)
        def on_key_change(*_): _save_sort_cfg()
        key_var.trace_add("write", on_key_change)
        def remove_this_row():
            idx = sort_key_rows.index(row_rec)
            row_rec["frame"].destroy()
            sort_key_rows.pop(idx)
            for i, rec in enumerate(sort_key_rows):
                for child in rec["frame"].winfo_children():
                    if isinstance(child, tb.Label) and child.cget("text").startswith("Key "):
                        child.configure(text=f"Key {i+1}:")
                rec["frame"].grid(row=i, column=0, sticky="we", pady=2)
            _save_sort_cfg()
        tb.Button(frame, text="Remove", command=remove_this_row).grid(row=0, column=3, padx=(6, 0))
        row_rec = {"frame": frame, "key_var": key_var, "desc_var": desc_var, "key_combo": key_combo}
        sort_key_rows.append(row_rec)
        path = sort_input_file_var.get()
        if path and os.path.exists(path):
            try:
                keys = _extract_keys_for_dropdown(path)
                key_combo["values"] = keys
            except Exception:
                pass
        _save_sort_cfg()

    btns_frame = tb.Frame(sort_tab)
    btns_frame.grid(row=6, column=0, columnspan=3, sticky="w", padx=5, pady=(6, 0))
    tb.Button(btns_frame, text="Add Key", command=lambda: add_sort_key_row()).grid(row=0, column=0, padx=(0, 6))
    def clear_keys():
        for rec in list(sort_key_rows):
            rec["frame"].destroy()
        sort_key_rows.clear(); _save_sort_cfg()
    tb.Button(btns_frame, text="Clear Keys", command=clear_keys).grid(row=0, column=1)

    _saved_specs = cfg.get("sort_specs", [])
    if isinstance(_saved_specs, list) and _saved_specs:
        for spec in _saved_specs:
            add_sort_key_row(prefill_key=str(spec.get("key", "")), prefill_desc=bool(spec.get("desc", False)))
    else:
        add_sort_key_row()

    tb.Button(sort_tab, text="Run Sort", bootstyle="info", command=run_sort)\
        .grid(row=7, column=0, columnspan=3, pady=10, sticky="w")

    sort_log_frame = tb.Labelframe(sort_tab, text="Sort Log")
    sort_log_frame.grid(row=8, column=0, columnspan=3, sticky="nsew", padx=5, pady=5)
    sort_log = ScrolledText(sort_log_frame, height=10, state="disabled")
    sort_log.pack(fill="both", expand=True, padx=5, pady=5)

    def _refresh_key_choices_if_possible(*_):
        path = sort_input_file_var.get()
        if not (path and os.path.exists(path)): return
        try:
            keys = _extract_keys_for_dropdown(path)
            for rec in sort_key_rows:
                rec["key_combo"]["values"] = keys
        except Exception:
            pass
        _save_sort_cfg()
    sort_input_file_var.trace_add("write", _refresh_key_choices_if_possible)
    sort_output_file_var.trace_add("write", lambda *_: _save_sort_cfg())

    # ---------------- Excel → JSON Tab ----------------
    excel_tab = tb.Frame(nb)
    nb.add(excel_tab, text="Excel → JSON")
    excel_tab.columnconfigure(1, weight=1)
    excel_tab.rowconfigure(7, weight=1)

    excel_input_file_var  = tk.StringVar(value=cfg.get("excel_input_file", ""))
    excel_sheet_var       = tk.StringVar(value=cfg.get("excel_sheet", ""))
    excel_header_row_var  = tk.BooleanVar(value=bool(cfg.get("excel_header_row", True)))
    excel_output_path_var = tk.StringVar(value=cfg.get("excel_output_path", ""))
    excel_orient_var      = tk.StringVar(value=cfg.get("excel_orient", "records"))
    excel_jsonl_var       = tk.BooleanVar(value=bool(cfg.get("excel_jsonl", False)))
    excel_indent_var      = tk.StringVar(value=str(cfg.get("excel_indent", "2")))
    excel_ascii_var       = tk.BooleanVar(value=bool(cfg.get("excel_ascii", False)))

    def _save_excel_cfg():
        indent_raw = (excel_indent_var.get() or "").strip()
        indent_val = int(indent_raw) if indent_raw.isdigit() else 2
        cfg.update({
            "excel_input_file":  excel_input_file_var.get(),
            "excel_sheet":       excel_sheet_var.get(),
            "excel_header_row":  bool(excel_header_row_var.get()),
            "excel_output_path": excel_output_path_var.get(),
            "excel_orient":      excel_orient_var.get(),
            "excel_jsonl":       bool(excel_jsonl_var.get()),
            "excel_indent":      indent_val,
            "excel_ascii":       bool(excel_ascii_var.get()),
        })
        save_config(cfg)

    def _excel_list_sheets(path: str) -> list[str]:
        if not path or not os.path.exists(path): return []
        try:
            import pandas as pd
            with pd.ExcelFile(path) as xf:
                return list(xf.sheet_names)
        except Exception as e:
            messagebox.showerror("Excel Error", f"Could not read sheets from:\n{path}\n\n{e}\n\nTip: pip install pandas openpyxl")
            return []

    def browse_excel_file(var):
        path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xlsm *.xls *.xlsb"), ("All files", "*.*")])
        if not path: return
        var.set(path)
        sheets = _excel_list_sheets(path)
        excel_sheet_combo.configure(values=sheets)
        if sheets:
            prev = cfg.get("excel_sheet")
            excel_sheet_var.set(prev if prev in sheets else sheets[0])
        _save_excel_cfg()

    r = 0
    tb.Label(excel_tab, text="Input Excel File:").grid(row=r, column=0, sticky="w")
    tb.Entry(excel_tab, textvariable=excel_input_file_var, state="readonly").grid(row=r, column=1, sticky="we", padx=5)
    tb.Button(excel_tab, text="Browse", command=lambda: browse_excel_file(excel_input_file_var)).grid(row=r, column=2); r+=1

    tb.Label(excel_tab, text="Sheet:").grid(row=r, column=0, sticky="w")
    excel_sheet_combo = tb.Combobox(excel_tab, textvariable=excel_sheet_var, state="readonly")
    excel_sheet_combo.grid(row=r, column=1, sticky="we", padx=5)
    tb.Button(excel_tab, text="Refresh", command=lambda: (
        excel_sheet_combo.configure(values=_excel_list_sheets(excel_input_file_var.get())), _save_excel_cfg()
    )).grid(row=r, column=2); r+=1

    tb.Checkbutton(excel_tab, text="First row contains headers", variable=excel_header_row_var, command=_save_excel_cfg)\
        .grid(row=r, column=0, sticky="w", padx=5); r+=1

    tb.Label(excel_tab, text="Output File (.json or .jsonl):").grid(row=r, column=0, sticky="w")
    tb.Entry(excel_tab, textvariable=excel_output_path_var, state="readonly").grid(row=r, column=1, sticky="we", padx=5)
    tb.Button(excel_tab, text="Save As",
              command=lambda: browse_json_or_jsonl_save(excel_output_path_var, use_jsonl_var=excel_jsonl_var)).grid(row=r, column=2); r+=1

    opts = tb.Labelframe(excel_tab, text="JSON Options")
    opts.grid(row=r, column=0, columnspan=3, sticky="nsew", padx=5, pady=(6, 4))
    opts.columnconfigure(1, weight=1)

    tb.Label(opts, text="Orient:").grid(row=0, column=0, sticky="w")
    tb.Combobox(opts, textvariable=excel_orient_var, state="readonly",
                values=["records", "table", "split", "index", "columns", "values"], width=12)\
        .grid(row=0, column=1, sticky="w", padx=5)
    tb.Checkbutton(opts, text="JSON Lines (.jsonl)", variable=excel_jsonl_var, command=_save_excel_cfg)\
        .grid(row=0, column=2, sticky="w", padx=10)

    tb.Label(opts, text="Indent (pretty-print):").grid(row=1, column=0, sticky="w", pady=(6,0))
    tb.Entry(opts, textvariable=excel_indent_var, width=8).grid(row=1, column=1, sticky="w", padx=5, pady=(6,0))
    tb.Checkbutton(opts, text="Force ASCII (escape non-ASCII)", variable=excel_ascii_var, command=_save_excel_cfg)\
        .grid(row=1, column=2, sticky="w", padx=10, pady=(6,0)); r+=1

    convert_btn = tb.Button(excel_tab, text="Convert Excel → JSON", bootstyle="primary")
    convert_btn.grid(row=r, column=0, columnspan=3, pady=10, sticky="w"); r+=1

    excel_log_frame = tb.Labelframe(excel_tab, text="Excel → JSON Log")
    excel_log_frame.grid(row=r, column=0, columnspan=3, sticky="nsew", padx=5, pady=5)
    excel_tab.rowconfigure(r, weight=1)
    excel_log = ScrolledText(excel_log_frame, height=10, state="disabled")
    excel_log.pack(fill="both", expand=True, padx=5, pady=5); r+=1

    if excel_input_file_var.get() and os.path.exists(excel_input_file_var.get()):
        try:
            sheets = _excel_list_sheets(excel_input_file_var.get())
            excel_sheet_combo.configure(values=sheets)
            if sheets and excel_sheet_var.get() not in sheets:
                excel_sheet_var.set(sheets[0])
        except Exception:
            pass

    def run_excel_to_json():
        in_path = (excel_input_file_var.get() or "").strip()
        out_path = (excel_output_path_var.get() or "").strip()
        sheet   = (excel_sheet_var.get() or "").strip()
        if not in_path: messagebox.showwarning("Missing Input", "Please choose an Excel file."); return
        if not out_path: messagebox.showwarning("Missing Output", "Please choose an output .json or .jsonl file."); return
        if not sheet: messagebox.showwarning("Missing Sheet", "Please choose a worksheet."); return
        _save_excel_cfg(); _set_active_log(excel_log)

        # You can tune this default. 1000 is a good general-purpose batch size.
        STREAM_BATCH_SIZE = 1000

        def _work():
            print("[excel→json] Starting…\n")
            try:
                import pandas as pd
            except Exception:
                print("pandas is required. Install with: pip install pandas openpyxl\n"); raise

            try:
                header_opt = 0 if bool(excel_header_row_var.get()) else None

                # If headers are present, coerce UPC to text with leading zeros preserved
                converters = None
                if header_opt == 0:
                    # Accept either 'upc' or 'UPC'
                    converters = {"upc": _excel_upc_to_text, "UPC": _excel_upc_to_text}

                # Read Excel
                try:
                    df = pd.read_excel(
                        in_path,
                        sheet_name=sheet,
                        header=header_opt,
                        engine=None,
                        converters=converters
                    )
                except Exception:
                    df = pd.read_excel(
                        in_path,
                        sheet_name=sheet,
                        header=header_opt,
                        engine="openpyxl",
                        converters=converters
                    )

                # If no header row, synthesize column names
                if header_opt is None:
                    df.columns = [f"col_{i+1}" for i in range(df.shape[1])]

                orient  = excel_orient_var.get()
                lines   = bool(excel_jsonl_var.get())
                ascii_f = bool(excel_ascii_var.get())
                indent_raw = (excel_indent_var.get() or "").strip()
                indent = int(indent_raw) if (indent_raw.isdigit() and not lines) else None

                # Validate orient for streaming:
                # - JSON Lines: we always write records-shaped objects per line
                # - JSON Array: we stream an array of record objects
                if orient not in ("records", "table", "split", "index", "columns", "values"):
                    raise ValueError(f"Unsupported orient: {orient}")

                # For streaming, we will treat everything as "records" per object.
                # If user chose non-records orient, we fall back to records (streamable)
                if orient != "records":
                    print(f"[excel→json] Note: Streaming currently supported as 'records' only; overriding orient='{orient}' → 'records'.")
                    orient = "records"

                total_rows = len(df)
                total_cols = len(df.columns)

                # Build a batch iterator of record dicts
                records_iter = _iter_df_records_in_batches(df, STREAM_BATCH_SIZE)

                # Stream to file atomically
                _stream_json_records_atomic(
                    path=out_path,
                    records_iter=records_iter,
                    ndjson=lines,           # True → JSONL, False → JSON array
                    indent=indent,
                    force_ascii=ascii_f
                )

                print(f"Wrote {out_path}\nRows: {total_rows}  Columns: {total_cols}")
                print("\n[excel→json] Completed successfully.\n")
            except Exception as e:
                print(f"[error] {e}\n")

        _run_in_thread(_work)

    convert_btn.configure(command=run_excel_to_json)

    for v in (excel_input_file_var, excel_sheet_var, excel_output_path_var, excel_orient_var, excel_indent_var):
        v.trace_add("write", lambda *_: _save_excel_cfg())

    # ---------------- JSON → Excel Tab ----------------
    j2x_tab = tb.Frame(nb)
    nb.add(j2x_tab, text="JSON → Excel")
    j2x_tab.columnconfigure(1, weight=1)
    j2x_tab.rowconfigure(6, weight=1)

    j2x_input_file_var  = tk.StringVar(value=cfg.get("j2x_input_file", ""))
    j2x_output_file_var = tk.StringVar(value=cfg.get("j2x_output_file", ""))
    j2x_sheet_name_var  = tk.StringVar(value=cfg.get("j2x_sheet_name", "Sheet1"))
    j2x_ndjson_var      = tk.BooleanVar(value=bool(cfg.get("j2x_ndjson", False)))
    j2x_flatten_var     = tk.BooleanVar(value=bool(cfg.get("j2x_flatten", True)))
    j2x_index_var       = tk.BooleanVar(value=bool(cfg.get("j2x_index", False)))

    def _save_j2x_cfg():
        cfg.update({
            "j2x_input_file":  j2x_input_file_var.get(),
            "j2x_output_file": j2x_output_file_var.get(),
            "j2x_sheet_name":  j2x_sheet_name_var.get(),
            "j2x_ndjson":      bool(j2x_ndjson_var.get()),
            "j2x_flatten":     bool(j2x_flatten_var.get()),
            "j2x_index":       bool(j2x_index_var.get()),
        }); save_config(cfg)

    r = 0
    tb.Label(j2x_tab, text="Input JSON/NDJSON File:").grid(row=r, column=0, sticky="w")
    tb.Entry(j2x_tab, textvariable=j2x_input_file_var, state="readonly").grid(row=r, column=1, sticky="we", padx=5)
    tb.Button(j2x_tab, text="Browse", command=lambda: _j2x_browse_json_input(j2x_input_file_var)).grid(row=r, column=2); r+=1

    tb.Checkbutton(j2x_tab, text="Input is NDJSON (one JSON object per line)", variable=j2x_ndjson_var, command=_save_j2x_cfg)\
        .grid(row=r, column=0, columnspan=3, sticky="w", padx=5, pady=(6, 0)); r+=1

    tb.Label(j2x_tab, text="Output Excel File (.xlsx):").grid(row=r, column=0, sticky="w")
    tb.Entry(j2x_tab, textvariable=j2x_output_file_var, state="readonly").grid(row=r, column=1, sticky="we", padx=5)
    tb.Button(j2x_tab, text="Save As", command=lambda: _j2x_browse_excel_save(j2x_output_file_var)).grid(row=r, column=2); r+=1

    j2x_opts = tb.Labelframe(j2x_tab, text="Excel Options")
    j2x_opts.grid(row=r, column=0, columnspan=3, sticky="nsew", padx=5, pady=(6, 4))
    j2x_opts.columnconfigure(1, weight=1)
    tb.Label(j2x_opts, text="Sheet name:").grid(row=0, column=0, sticky="w")
    tb.Entry(j2x_opts, textvariable=j2x_sheet_name_var, width=24).grid(row=0, column=1, sticky="w", padx=5)
    tb.Checkbutton(j2x_opts, text="Flatten nested objects (json_normalize)", variable=j2x_flatten_var, command=_save_j2x_cfg)\
        .grid(row=0, column=2, sticky="w", padx=10)
    tb.Checkbutton(j2x_opts, text="Include index column", variable=j2x_index_var, command=_save_j2x_cfg)\
        .grid(row=1, column=2, sticky="w", padx=10, pady=(6, 0)); r+=1

    j2x_run_btn = tb.Button(j2x_tab, text="Convert JSON → Excel", bootstyle="primary")
    j2x_run_btn.grid(row=r, column=0, columnspan=3, pady=10, sticky="w"); r+=1

    j2x_log_frame = tb.Labelframe(j2x_tab, text="JSON → Excel Log")
    j2x_log_frame.grid(row=r, column=0, columnspan=3, sticky="nsew", padx=5, pady=5)
    j2x_tab.rowconfigure(r, weight=1)
    j2x_log = ScrolledText(j2x_log_frame, height=10, state="disabled")
    j2x_log.pack(fill="both", expand=True, padx=5, pady=5); r+=1

    def run_json_to_excel():
        in_path  = (j2x_input_file_var.get() or "").strip()
        out_path = (j2x_output_file_var.get() or "").strip()
        sheet    = (j2x_sheet_name_var.get() or "").strip() or "Sheet1"
        ndjson   = bool(j2x_ndjson_var.get())
        flatten  = bool(j2x_flatten_var.get())
        inc_idx  = bool(j2x_index_var.get())
        if not in_path: messagebox.showwarning("Missing Input", "Please choose a JSON or NDJSON file."); return
        if not out_path: messagebox.showwarning("Missing Output", "Please choose an output .xlsx file."); return
        _save_j2x_cfg(); _set_active_log(j2x_log)
        def _work():
            print("[json→excel] Starting…\n")
            try:
                import pandas as pd
            except Exception:
                print("pandas is required. Install with: pip install pandas openpyxl\n"); raise
            try:
                records = []
                if ndjson:
                    try: text = _read_file(in_path)
                    except Exception:
                        with open(in_path, "r", encoding="utf-8") as f: text = f.read()
                    text = _fix_nan_tokens_local(text)
                    for i, line in enumerate(text.splitlines(), 1):
                        line = line.strip()
                        if not line: continue
                        try: records.append(json.loads(line))
                        except Exception as e:
                            print(f"Warning: skipped NDJSON line {i}: {e}")
                else:
                    try: raw = _read_file(in_path)
                    except Exception:
                        with open(in_path, "r", encoding="utf-8") as f: raw = f.read()
                    raw = _fix_nan_tokens_local(raw)
                    data = json.loads(raw)
                    if isinstance(data, list): records = data
                    elif isinstance(data, dict): records = [data]
                    else: raise ValueError("Top-level JSON must be an array or object.")
                if flatten:
                    df = pd.json_normalize(records, sep=".")
                else:
                    df = pd.DataFrame(records)
                try:
                    from openpyxl.utils import get_column_letter
                except Exception:
                    print("Note: openpyxl not found — install 'openpyxl' for auto column width.\n")
                with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
                    df.to_excel(writer, sheet_name=sheet, index=inc_idx)
                    try:
                        ws = writer.sheets[sheet]
                        from openpyxl.utils import get_column_letter
                        for idx, col in enumerate(df.columns, 1):
                            header = str(col) if col is not None else ""
                            max_len = len(header)
                            for v in df[col].head(100).astype(str):
                                if len(v) > max_len: max_len = len(v)
                            width = min(max_len + 2, 60)
                            ws.column_dimensions[get_column_letter(idx + (1 if inc_idx else 0))].width = width
                    except Exception:
                        pass
                print(f"Wrote {out_path}\nRows: {len(df)}  Columns: {len(df.columns)}  Sheet: {sheet}")
                print("\n[json→excel] Completed successfully.\n")
            except Exception as e:
                print(f"[error] {e}\n")
        _run_in_thread(_work)
    j2x_run_btn.configure(command=run_json_to_excel)

    for v in (j2x_input_file_var, j2x_output_file_var, j2x_sheet_name_var):
        v.trace_add("write", lambda *_: _save_j2x_cfg())

    # ---------------- Patch Columns Tab (NEW) ----------------

    # ---------- Reusable checkbox list with Select All + scroll ----------
    class CheckListPanel(tb.Labelframe):
        def __init__(self, master, title: str):
            super().__init__(master, text=title)
            self.columnconfigure(0, weight=1)
            # Top bar with Select All
            top = tb.Frame(self)
            top.grid(row=0, column=0, sticky="we", padx=5, pady=(5, 2))
            top.columnconfigure(0, weight=1)

            self._select_all_var = tk.BooleanVar(value=False)
            self._select_all = tb.Checkbutton(top, text="Select All", variable=self._select_all_var,
                                            command=self._on_select_all)
            self._select_all.grid(row=0, column=0, sticky="w")

            # Scrollable area
            holder = tb.Frame(self)
            holder.grid(row=1, column=0, sticky="nsew")
            self.rowconfigure(1, weight=1)

            self._canvas = tk.Canvas(holder, highlightthickness=0)
            self._vsb = tb.Scrollbar(holder, orient="vertical", command=self._canvas.yview)
            self._list_frame = tb.Frame(self._canvas)

            self._list_frame.bind("<Configure>",
                                lambda e: self._canvas.configure(scrollregion=self._canvas.bbox("all")))
            self._canvas.create_window((0, 0), window=self._list_frame, anchor="nw")
            self._canvas.configure(yscrollcommand=self._vsb.set)

            self._canvas.pack(side="left", fill="both", expand=True)
            self._vsb.pack(side="right", fill="y")

            # storage
            self._items: list[str] = []
            self._vars: dict[str, tk.BooleanVar] = {}

        def set_items(self, items: list[str], prechecked: set[str] | None = None):
            """Replace list items; precheck those in 'prechecked'."""
            prechecked = prechecked or set()
            # clear old
            for w in self._list_frame.winfo_children():
                w.destroy()
            self._items = list(items)
            self._vars.clear()
            # build
            for i, name in enumerate(self._items):
                var = tk.BooleanVar(value=(name in prechecked))
                cb = tb.Checkbutton(self._list_frame, text=name, variable=var,
                                    command=self._on_any_item_toggle)
                cb.grid(row=i, column=0, sticky="w", padx=6, pady=2)
                self._vars[name] = var
            # update select all state
            self._sync_select_all()

        def get_selected(self) -> list[str]:
            return [k for k, v in self._vars.items() if bool(v.get())]

        def _on_select_all(self):
            state = bool(self._select_all_var.get())
            for v in self._vars.values():
                v.set(state)

        def _on_any_item_toggle(self):
            # if any unchecked -> select_all off; if all checked -> on
            self._sync_select_all()

        def _sync_select_all(self):
            vals = [bool(v.get()) for v in self._vars.values()]
            self._select_all_var.set(bool(vals) and all(vals))

    def _collect_top_level_keys_from_json_text(text: str, limit: int = 400) -> set[str]:
        """Collect top-level keys from JSON array (sample) or NDJSON text."""
        keys = set()
        fixed = _fix_nan_tokens(text)
        # Try as standard JSON
        try:
            data = json.loads(fixed)
            if isinstance(data, list):
                sample = data[:min(len(data), 400)]
                for obj in sample:
                    if isinstance(obj, dict):
                        keys.update(obj.keys())
                return keys
            if isinstance(data, dict):
                keys.update(data.keys())
                return keys
        except Exception:
            pass
        # Fallback: NDJSON
        for i, line in enumerate(fixed.splitlines(), 1):
            if i > limit:
                break
            line = line.strip()
            if not line:
                continue
            try:
                obj = json.loads(line)
                if isinstance(obj, dict):
                    keys.update(obj.keys())
            except Exception:
                continue
        return keys


    def _read_json_text(path: str) -> str:
        try:
            return _read_file(path)
        except Exception:
            with open(path, "r", encoding="utf-8") as f:
                return f.read()



    # ---------------- Patch Columns Tab UI ----------------
    patch_tab = tb.Frame(nb)
    nb.add(patch_tab, text="Patch Columns")
    patch_tab.columnconfigure(1, weight=1)
    # rows that will expand:
    # - lists frame (row 8)
    # - log frame (row 11)
    patch_tab.rowconfigure(8, weight=2)
    patch_tab.rowconfigure(11, weight=1)

    # --- State vars (persisted) ---
    patch_original_file_var = tk.StringVar(value=cfg.get("patch_original_file", ""))
    patch_updated_file_var  = tk.StringVar(value=cfg.get("patch_updated_file", ""))
    patch_output_file_var   = tk.StringVar(value=cfg.get("patch_output_file", ""))

    patch_id_field_var      = tk.StringVar(value=cfg.get("patch_id_field", "item_#"))
    patch_ndjson_var        = tk.BooleanVar(value=bool(cfg.get("patch_ndjson", False)))
    patch_fixnan_var        = tk.BooleanVar(value=bool(cfg.get("patch_fixnan", True)))
    patch_append_new_var    = tk.BooleanVar(value=bool(cfg.get("patch_append_new", False)))

    # selections persisted
    patch_overwrite_sel_cfg = set(cfg.get("patch_overwrite_selected", []))
    patch_add_sel_cfg       = set(cfg.get("patch_add_selected", []))

    # --- Persist helper ---
    def _save_patch_cfg():
        cfg.update({
            "patch_original_file": patch_original_file_var.get(),
            "patch_updated_file":  patch_updated_file_var.get(),
            "patch_output_file":   patch_output_file_var.get(),
            "patch_id_field":      patch_id_field_var.get(),
            "patch_ndjson":        bool(patch_ndjson_var.get()),
            "patch_fixnan":        bool(patch_fixnan_var.get()),
            "patch_append_new":    bool(patch_append_new_var.get()),
            "patch_overwrite_selected": overwrite_panel.get_selected(),
            "patch_add_selected":       add_panel.get_selected(),
        })
        save_config(cfg)

    # --- Browsers ---
    def _browse_json_input(var):
        path = filedialog.askopenfilename(
            filetypes=[("JSON/NDJSON files", "*.json *.ndjson"), ("All files", "*.*")]
        )
        if path:
            var.set(path)
            _save_patch_cfg()
            _refresh_patch_keys()

    def _browse_json_output(var):
        default_ext = ".ndjson" if patch_ndjson_var.get() else ".json"
        path = filedialog.asksaveasfilename(
            defaultextension=default_ext,
            filetypes=[("JSON", "*.json"), ("NDJSON", "*.ndjson"), ("All files", "*.*")]
        )
        if path:
            var.set(path)
            _save_patch_cfg()

    # --- Layout ---
    r = 0
    tb.Label(patch_tab, text="Original JSON/NDJSON File:").grid(row=r, column=0, sticky="w")
    tb.Entry(patch_tab, textvariable=patch_original_file_var, state="readonly")\
        .grid(row=r, column=1, sticky="we", padx=5)
    tb.Button(patch_tab, text="Browse", command=lambda: _browse_json_input(patch_original_file_var))\
        .grid(row=r, column=2)
    r += 1

    tb.Label(patch_tab, text="Updated JSON/NDJSON File:").grid(row=r, column=0, sticky="w")
    tb.Entry(patch_tab, textvariable=patch_updated_file_var, state="readonly")\
        .grid(row=r, column=1, sticky="we", padx=5)
    tb.Button(patch_tab, text="Browse", command=lambda: _browse_json_input(patch_updated_file_var))\
        .grid(row=r, column=2)
    r += 1

    tb.Label(patch_tab, text="Output File:").grid(row=r, column=0, sticky="w")
    tb.Entry(patch_tab, textvariable=patch_output_file_var, state="readonly")\
        .grid(row=r, column=1, sticky="we", padx=5)
    tb.Button(patch_tab, text="Save As", command=lambda: _browse_json_output(patch_output_file_var))\
        .grid(row=r, column=2)
    r += 1

    tb.Label(patch_tab, text="ID Field:").grid(row=r, column=0, sticky="w")
    patch_id_combo = tb.Combobox(patch_tab, textvariable=patch_id_field_var, state="normal")  # allows typing
    patch_id_combo.grid(row=r, column=1, sticky="w", padx=5)
    tb.Button(patch_tab, text="Refresh Keys",
            command=lambda: (_refresh_patch_keys(), _save_patch_cfg()))\
    .grid(row=r, column=2)
    r += 1

    tb.Checkbutton(patch_tab, text="Input/Output are NDJSON", variable=patch_ndjson_var,
                command=lambda: (_save_patch_cfg()))\
    .grid(row=r, column=0, sticky="w", padx=5, pady=(6, 0))
    r += 1

    tb.Checkbutton(patch_tab, text="Fix bare NaN → null", variable=patch_fixnan_var,
                command=lambda: (_save_patch_cfg()))\
    .grid(row=r, column=0, sticky="w", padx=5)
    r += 1

    tb.Checkbutton(patch_tab, text="Append records present only in Updated", variable=patch_append_new_var,
                command=lambda: (_save_patch_cfg()))\
    .grid(row=r, column=0, sticky="w", padx=5, pady=(0, 6))
    r += 1

    # --- Two side-by-side checklists (expandable row) ---
    lists_frame = tb.Frame(patch_tab)
    lists_frame.grid(row=r, column=0, columnspan=3, sticky="nsew", padx=5, pady=(0, 6))
    patch_tab.rowconfigure(r, weight=2)
    lists_frame.columnconfigure(0, weight=1)
    lists_frame.columnconfigure(1, weight=1)

    overwrite_panel = CheckListPanel(lists_frame, "Overwrite Columns (common in Original & Updated)")
    overwrite_panel.grid(row=0, column=0, sticky="nsew", padx=(0, 6))

    add_panel = CheckListPanel(lists_frame, "Add Columns (present in Updated only)")
    add_panel.grid(row=0, column=1, sticky="nsew", padx=(6, 0))
    r += 1

    # --- Run button ---
    run_patch_btn = tb.Button(patch_tab, text="Run Patch", bootstyle="warning")
    run_patch_btn.grid(row=r, column=0, columnspan=3, pady=8, sticky="w")
    r += 1

    # --- Log (expandable) ---
    patch_log_frame = tb.Labelframe(patch_tab, text="Patch Log")
    patch_log_frame.grid(row=r, column=0, columnspan=3, sticky="nsew", padx=5, pady=5)
    patch_tab.rowconfigure(r, weight=1)
    patch_log = ScrolledText(patch_log_frame, height=10, state="disabled")
    patch_log.pack(fill="both", expand=True, padx=5, pady=5)
    r += 1

    # --- Populate ID key choices + lists from files ---
    def _refresh_patch_keys():
        orig_path = (patch_original_file_var.get() or "").strip()
        upd_path  = (patch_updated_file_var.get() or "").strip()
        orig_keys, upd_keys = set(), set()

        if orig_path and os.path.exists(orig_path):
            try:
                orig_text = _read_json_text(orig_path)
                orig_keys = _collect_top_level_keys_from_json_text(orig_text)
            except Exception:
                orig_keys = set()

        if upd_path and os.path.exists(upd_path):
            try:
                upd_text = _read_json_text(upd_path)
                upd_keys = _collect_top_level_keys_from_json_text(upd_text)
            except Exception:
                upd_keys = set()

        # ID combo: prefer previous saved value if in either set
        id_vals = sorted(orig_keys | upd_keys)
        patch_id_combo.configure(values=id_vals)
        if patch_id_field_var.get() not in id_vals and id_vals:
            patch_id_field_var.set(id_vals[0])

        # Fill lists
        common = sorted(orig_keys & upd_keys)
        added  = sorted(upd_keys - orig_keys)

        # reapply saved selections, but only those still present
        overwrite_panel.set_items(common, prechecked=(patch_overwrite_sel_cfg & set(common)))
        add_panel.set_items(added, prechecked=(patch_add_sel_cfg & set(added)))

    # initialize lists if files are already chosen
    _refresh_patch_keys()

    # --- Worker: run patch ---
    def _load_items_for_patch(path: str, ndjson: bool, fixnan: bool) -> list[dict]:
        if not path:
            return []
        if ndjson:
            text = _read_json_text(path)
            if fixnan:
                text = _fix_nan_tokens(text)
            out = []
            for i, line in enumerate(text.splitlines(), 1):
                line = line.strip()
                if not line:
                    continue
                try:
                    obj = json.loads(line)
                    if isinstance(obj, dict):
                        out.append(obj)
                    else:
                        raise ValueError("NDJSON line is not an object")
                except Exception as e:
                    raise ValueError(f"NDJSON parse error line {i}: {e}") from e
            return out
        else:
            text = _read_json_text(path)
            if fixnan:
                text = _fix_nan_tokens(text)
            data = json.loads(text)
            if isinstance(data, list):
                return [x for x in data if isinstance(x, dict)]
            if isinstance(data, dict):
                return [data]
            raise ValueError("JSON must be array or object of records")

    def run_patch():
        orig_path = (patch_original_file_var.get() or "").strip()
        upd_path  = (patch_updated_file_var.get() or "").strip()
        out_path  = (patch_output_file_var.get() or "").strip()
        id_field  = (patch_id_field_var.get() or "").strip()

        if not orig_path:
            messagebox.showwarning("Missing Original", "Please choose the ORIGINAL JSON/NDJSON file.")
            return
        if not upd_path:
            messagebox.showwarning("Missing Updated", "Please choose the UPDATED JSON/NDJSON file.")
            return
        if not out_path:
            messagebox.showwarning("Missing Output", "Please choose an OUTPUT file.")
            return
        if not id_field:
            messagebox.showwarning("Missing ID", "Please select or enter the ID field.")
            return

        _save_patch_cfg()
        _set_active_log(patch_log)

        ndjson  = bool(patch_ndjson_var.get())
        fixnan  = bool(patch_fixnan_var.get())
        append_new = bool(patch_append_new_var.get())

        overwrite_cols = overwrite_panel.get_selected()
        add_cols       = add_panel.get_selected()

        def _work():
            print("[patch] Starting…\n")
            try:
                # load
                original = _load_items_for_patch(orig_path, ndjson, fixnan)
                updated  = _load_items_for_patch(upd_path,  ndjson, fixnan)

                if not original:
                    raise ValueError("Original file has no records.")

                # index by ID
                orig_idx: dict[str, dict] = {}
                order: list[str] = []
                missing_id_orig = 0
                for i, obj in enumerate(original):
                    key = obj.get(id_field)
                    if key in (None, ""):
                        missing_id_orig += 1
                        continue
                    key = str(key)
                    if key not in orig_idx:
                        order.append(key)
                    orig_idx[key] = obj

                upd_idx: dict[str, dict] = {}
                missing_id_upd = 0
                for obj in updated:
                    key = obj.get(id_field)
                    if key in (None, ""):
                        missing_id_upd += 1
                        continue
                    upd_idx[str(key)] = obj

                if missing_id_orig:
                    print(f"Warning: {missing_id_orig} original record(s) missing '{id_field}' were ignored.")
                if missing_id_upd:
                    print(f"Warning: {missing_id_upd} updated record(s) missing '{id_field}' were ignored.")

                # patch
                patched: list[dict] = []
                overwrites_applied = 0
                adds_applied = 0
                updated_found = 0
                not_found = 0

                for key in order:
                    base = dict(orig_idx[key])  # copy
                    upd = upd_idx.get(key)
                    if upd is None:
                        patched.append(base)
                        not_found += 1
                        continue

                    updated_found += 1

                    # Overwrite: only for fields selected; require field to exist in both? (we computed list as common)
                    for col in overwrite_cols:
                        if col in upd and col in base:
                            if base.get(col) != upd.get(col):
                                base[col] = upd[col]
                                overwrites_applied += 1

                    # Add: only if field absent in base, present in upd
                    for col in add_cols:
                        if col not in base and col in upd:
                            base[col] = upd[col]
                            adds_applied += 1

                    patched.append(base)

                # Append new records present only in updated
                extras = [upd for k, upd in upd_idx.items() if k not in orig_idx]
                if append_new and extras:
                    patched.extend(extras)

                # write
                if ndjson:
                    payload = "\n".join(json.dumps(o, ensure_ascii=False) for o in patched) + "\n"
                    _write_atomic(out_path, payload)
                else:
                    _write_atomic(out_path, json.dumps(patched, ensure_ascii=False, indent=2))

                print(f"Patched → {out_path}")
                print(f"Original records        : {len(original)}")
                print(f"Updated records         : {len(updated)}")
                print(f"Matched by ID           : {updated_found}")
                print(f"Not found in updated    : {not_found}")
                print(f"Overwrites applied      : {overwrites_applied}")
                print(f"Adds applied            : {adds_applied}")
                print(f"Appended new records    : {len(extras) if append_new else 0} (append={'ON' if append_new else 'OFF'})")
                print("\n[patch] Completed successfully.\n")

            except Exception as e:
                print(f"[error] {e}\n")

        _run_in_thread(_work)

    run_patch_btn.configure(command=run_patch)

    # Persist when fields change
    for v in (patch_original_file_var, patch_updated_file_var, patch_output_file_var,
            patch_id_field_var):
        v.trace_add("write", lambda *_: _save_patch_cfg())


    # --- Remember last-opened tab ---
    def _on_tab_changed(event=None):
        try:
            idx = nb.index(nb.select())
            cfg["last_tab"] = int(idx); save_config(cfg)
        except Exception:
            pass
    nb.bind("<<NotebookTabChanged>>", _on_tab_changed)

    # Restore last-opened tab
    try:
        last_idx = int(cfg.get("last_tab", 0) or 0)
    except Exception:
        last_idx = 0
    tab_count = len(nb.tabs())
    nb.select(last_idx if 0 <= last_idx < tab_count else 0)

    # Close handler
    app.protocol("WM_DELETE_WINDOW", lambda: (_save_config_now(), app.destroy()))
    _drain_log_queue()
    app.mainloop()

# ---------------------------- Entrypoint ----------------------------
def _load_mapping_file(path: str | None) -> Dict[str, str] | None:
    if not path: return None
    try:
        data = json.loads(_read_text(path))
        if isinstance(data, dict): return data
    except Exception:
        pass
    return None

def main():
    p = argparse.ArgumentParser(description="JSON Management Tool: Split/Merge/Group/Sort, Excel↔JSON, Patch Columns")
    sub = p.add_subparsers(dest="cmd")

    ps = sub.add_parser("split", help="Split a large JSON array (or NDJSON) into batches.")
    ps.add_argument("--input", required=True)
    ps.add_argument("--outdir", required=True)
    ps.add_argument("--size", type=int, default=50)
    ps.add_argument("--id-field", default="item_#")
    ps.add_argument("--ndjson", action="store_true")
    ps.add_argument("--fix-nan", action="store_true")

    pm = sub.add_parser("merge", help="Merge augmented batches back into original order using the ID field.")
    pm.add_argument("--original", required=True)
    pm.add_argument("--batches-dir", required=True)
    pm.add_argument("--output", required=True)
    pm.add_argument("--id-field", default="item_#")
    pm.add_argument("--ndjson", action="store_true")
    pm.add_argument("--fix-nan", action="store_true")

    pk = sub.add_parser("split-by", help="Split into files by (similar) values of a key.")
    pk.add_argument("--input", required=True)
    pk.add_argument("--outdir", required=True)
    pk.add_argument("--key", required=True)
    pk.add_argument("--mode", default="normalize",
                    help="exact|casefold|normalize|prefix:N|regex:<pat>|map:<file>")
    pk.add_argument("--min-bucket", type=int, default=0)
    pk.add_argument("--max-per-file", type=int, default=0)
    pk.add_argument("--missing-bucket", default="_missing")
    pk.add_argument("--ndjson", action="store_true")
    pk.add_argument("--fix-nan", action="store_true")

    # No CLI for sort/excel/jsonxl/patch; use GUI
    if len(sys.argv) == 1:
        launch_gui(); return

    args = p.parse_args()
    if not getattr(args, "cmd", None):
        launch_gui(); return

    if args.cmd == "split":
        cmd_split(args)
    elif args.cmd == "merge":
        cmd_merge(args)
    elif args.cmd == "split-by":
        cmd_split_by(args)
    else:
        p.print_help()

if __name__ == "__main__":
    main()
