"""
Excel generation module.

This module handles creating and formatting the output Excel spreadsheet.
"""

from pathlib import Path
from typing import List
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from src.models import ProductRecord
import src.config as config


class ExcelGenerator:
    """
    Handles Excel file creation and formatting.
    """

    def __init__(self):
        """Initialize the Excel generator."""
        self.font = Font(
            name=config.EXCEL_CONFIG["font_name"],
            size=config.EXCEL_CONFIG["font_size"]
        )

    def generate(self, records: List[ProductRecord], output_path: Path) -> None:
        """
        Generate Excel file from product records.

        Args:
            records: List of ProductRecord objects
            output_path: Path where Excel file will be saved
        """
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Products"

        # Write header row
        headers = config.EXCEL_CONFIG["columns"]
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = self.font

        # Write data rows
        for row_idx, record in enumerate(records, start=2):
            for col_idx, value in enumerate(record.to_list(), start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = self.font

        # Auto-fit column widths
        if config.EXCEL_CONFIG["auto_fit_columns"]:
            self._auto_fit_columns(ws)

        # Ensure output directory exists
        output_path.parent.mkdir(parents=True, exist_ok=True)

        # Save workbook
        wb.save(output_path)

    def _auto_fit_columns(self, worksheet) -> None:
        """
        Auto-fit column widths to content.

        Args:
            worksheet: openpyxl worksheet object
        """
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)

            for cell in column:
                try:
                    if cell.value:
                        # Calculate length of cell value
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                except:
                    pass

            # Set column width with some padding
            # Excel column width units are roughly character widths
            adjusted_width = max_length + 2
            worksheet.column_dimensions[column_letter].width = adjusted_width
