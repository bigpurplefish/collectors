"""
Aggregates Excel-to-Categorizer converter.

Reads product data from an Excel spreadsheet and converts it into
categorizer-compatible JSON format, grouping rows by parent_sku.
"""

import json
import os
import sys
import logging
from pathlib import Path
from collections import OrderedDict
from typing import Any, Callable, Dict, List, Optional

import openpyxl

# Import shared logging utilities
sys.path.insert(0, str(Path(__file__).parent.parent.parent / "shared"))
from utils.logging_utils import (
    log_and_status, log_section_header, log_progress,
    log_success, log_warning, log_error, log_summary,
)

# Expected Excel column order (0-indexed)
COLUMNS = [
    "sku", "parent_sku", "vendor", "title", "price", "cost",
    "color", "size", "unit_of_sale", "approximate_product_coverage",
    "image_file_name",
]

IMAGES_DIR = os.path.join(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
    "input", "images",
)


def _read_excel(input_file: str, status_fn: Optional[Callable]) -> List[Dict[str, Any]]:
    """Read Excel file and return list of row dicts."""
    log_and_status(status_fn, f"Reading Excel file: {input_file}")

    wb = openpyxl.load_workbook(input_file, read_only=True, data_only=True)
    ws = wb.active

    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    # Normalize header names
    headers = [str(h).strip().lower().replace(" ", "_") if h else f"col_{i}"
               for i, h in enumerate(headers)]

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_dict = {}
        for i, val in enumerate(row):
            key = headers[i] if i < len(headers) else f"col_{i}"
            row_dict[key] = val
        # Skip completely empty rows
        if any(v is not None for v in row_dict.values()):
            rows.append(row_dict)

    wb.close()
    log_and_status(status_fn, f"Read {len(rows)} data rows from Excel")
    return rows


def _group_by_parent_sku(rows: List[Dict]) -> OrderedDict:
    """Group rows by parent_sku, preserving insertion order."""
    groups: OrderedDict[Any, List[Dict]] = OrderedDict()
    for row in rows:
        parent = row.get("parent_sku")
        if parent is None:
            parent = row.get("sku")
        groups.setdefault(parent, []).append(row)
    return groups


def _find_parent_row(rows: List[Dict], parent_sku: Any) -> Dict:
    """Find the row where sku == parent_sku, or return the first row."""
    for row in rows:
        if row.get("sku") == parent_sku:
            return row
    return rows[0]


def _determine_options(rows: List[Dict]) -> List[Dict]:
    """Determine variant options for a group of rows.

    Returns list of option dicts, each with 'name', 'field', and 'values'.
    Empty list for standalone products (single row, no populated option fields).
    """
    if len(rows) <= 1:
        return []

    options = []

    # Check color
    colors = list(dict.fromkeys(
        str(c) for c in (r.get("color") for r in rows) if c is not None
    ))
    if colors:
        options.append({"name": "Color", "field": "color", "values": colors})

    # Check size
    sizes = list(dict.fromkeys(
        str(s) for s in (r.get("size") for r in rows) if s is not None
    ))
    if sizes:
        options.append({"name": "Size", "field": "size", "values": sizes})

    # Check unit_of_sale
    units = list(dict.fromkeys(
        str(u) for u in (r.get("unit_of_sale") for r in rows) if u is not None
    ))
    if units:
        options.append({"name": "Unit of Sale", "field": "unit_of_sale", "values": units})

    return options


def _format_price(value: Any) -> str:
    """Format a price value as string with 2 decimal places."""
    if value is None:
        return "0.00"
    try:
        return f"{float(value):.2f}"
    except (ValueError, TypeError):
        return "0.00"


def _build_description(parent_row: Dict, rows: List[Dict], options: List[Dict]) -> str:
    """Build a factual HTML description from available fields.

    The categorizer's AI will rewrite this into marketing copy.
    """
    title = parent_row.get("title", "Product")
    coverage = parent_row.get("approximate_product_coverage", "")

    parts = []

    # Product name and available variants
    option_descriptions = []
    for opt in options:
        if len(opt["values"]) > 1:
            vals = opt["values"]
            values_str = ", ".join(str(v) for v in vals[:-1])
            values_str += f", and {vals[-1]}"
            option_descriptions.append(f"{values_str} ({opt['name'].lower()} options)")
    if option_descriptions:
        parts.append(f"<p>{title} available in {'; '.join(option_descriptions)}.</p>")
    else:
        parts.append(f"<p>{title}.</p>")

    # Size info (only for products where size is not an option)
    option_fields = {opt["field"] for opt in options}
    size = parent_row.get("size")
    if size and "size" not in option_fields:
        parts.append(f"<p>Size: {size}.</p>")

    # Unit of sale and coverage (only mention unit if not an option)
    sale_coverage_parts = []
    unit_of_sale = parent_row.get("unit_of_sale", "")
    if unit_of_sale and "unit_of_sale" not in option_fields:
        sale_coverage_parts.append(f"Sold by the {unit_of_sale}")
    if coverage:
        sale_coverage_parts.append(f"Approximate coverage: {coverage}")
    if sale_coverage_parts:
        parts.append(f"<p>{'. '.join(sale_coverage_parts)}.</p>")

    return "".join(parts)


def _build_images(
    rows: List[Dict],
    s3_client,
    bucket: str,
    folder: str,
    status_fn: Optional[Callable],
) -> tuple:
    """Build images array and SKU-to-URL mapping.

    Returns (images_list, sku_to_image_url) where images_list has tagged alt text
    and sku_to_image_url maps each variant SKU to its uploaded image URL.
    """
    from src.s3_client import upload_image_to_s3

    images = []
    sku_to_image_url = {}
    seen_files = {}  # filename -> public_url

    for row in rows:
        image_file = row.get("image_file_name")
        if not image_file:
            continue

        sku = str(row.get("sku", ""))

        # If already uploaded, reuse URL
        if image_file in seen_files:
            if sku:
                sku_to_image_url[sku] = seen_files[image_file]
            continue

        abs_path = os.path.join(IMAGES_DIR, image_file)

        # Build hashtag alt text from option fields
        alt_parts = []
        color = row.get("color")
        if color:
            alt_parts.append(f"#{color}")
        size = row.get("size")
        if size:
            alt_parts.append(f"#{size}")
        unit = row.get("unit_of_sale")
        if unit:
            alt_parts.append(f"#{unit}")
        alt_text = " ".join(alt_parts) if alt_parts else row.get("title", "Product image")

        if not os.path.exists(abs_path):
            log_warning(
                status_fn,
                f"Image not found: {image_file}",
                details=f"Expected at {abs_path}",
            )
            continue

        try:
            public_url = upload_image_to_s3(
                s3_client, bucket, folder, abs_path, status_fn,
            )
            images.append({"src": public_url, "alt": alt_text})
            seen_files[image_file] = public_url
            if sku:
                sku_to_image_url[sku] = public_url
        except Exception as e:
            log_error(
                status_fn,
                f"Failed to upload {image_file}: {e}",
            )

    return images, sku_to_image_url


def _build_product(
    parent_sku: Any,
    rows: List[Dict],
    s3_client,
    bucket: str,
    folder: str,
    status_fn: Optional[Callable],
) -> Dict:
    """Build a single product dict from a group of rows."""
    parent_row = _find_parent_row(rows, parent_sku)
    options = _determine_options(rows)

    # Build images first to get sku-to-URL mapping for metafields
    images, sku_to_image_url = _build_images(rows, s3_client, bucket, folder, status_fn)

    product = {
        "title": parent_row.get("title", ""),
        "body_html": _build_description(parent_row, rows, options),
        "vendor": parent_row.get("vendor", ""),
        "status": "ACTIVE",
    }

    # Options
    if options:
        product["options"] = [
            {"name": opt["name"], "position": i + 1, "values": opt["values"]}
            for i, opt in enumerate(options)
        ]

    # Variants
    variants = []
    for row in rows:
        variant = {
            "sku": str(row.get("sku", "")),
            "price": _format_price(row.get("price")),
            "cost": _format_price(row.get("cost")),
            "inventoryQuantity": 0,
        }

        unit = str(row.get("unit_of_sale", "")).strip().lower()
        if unit == "ton":
            variant["weight"] = 2000.0
            variant["weight_unit"] = "lb"

        # Assign option values from options list
        for i, opt in enumerate(options):
            val = row.get(opt["field"])
            if val is not None:
                variant[f"option{i + 1}"] = str(val)

        # Add color_swatch_image metafield if variant has an uploaded image
        sku = str(row.get("sku", ""))
        if sku in sku_to_image_url:
            variant["metafields"] = [
                {"key": "color_swatch_image", "value": sku_to_image_url[sku]}
            ]

        variants.append(variant)

    product["variants"] = variants

    if images:
        product["images"] = images

    return product


def process_products(cfg: Dict[str, Any], status_fn: Optional[Callable] = None):
    """Main entry point: read Excel, convert to categorizer JSON, write output.

    Args:
        cfg: Configuration dictionary
        status_fn: Status callback for GUI/CLI updates
    """
    input_file = cfg.get("input_file", "")
    output_file = cfg.get("output_file", "")
    processing_mode = cfg.get("processing_mode", "skip")
    start_record = cfg.get("start_record", "")
    end_record = cfg.get("end_record", "")

    log_section_header(status_fn, "Aggregates Excel-to-Categorizer Converter")

    # Validate inputs
    if not input_file:
        log_error(status_fn, "No input file configured")
        return
    if not os.path.exists(input_file):
        log_error(status_fn, f"Input file not found: {input_file}")
        return
    if not output_file:
        log_error(status_fn, "No output file configured")
        return

    # Validate and initialize S3
    from src.s3_client import get_s3_client
    s3_bucket = cfg.get("S3_BUCKET", "").strip()
    s3_folder = cfg.get("S3_FOLDER", "").strip()
    if not cfg.get("AWS_ACCESS_KEY_ID") or not cfg.get("AWS_SECRET_ACCESS_KEY"):
        log_error(status_fn, "AWS credentials not configured. Open Settings to configure.")
        return
    if not s3_bucket:
        log_error(status_fn, "S3 Bucket not configured. Open Settings to configure.")
        return

    try:
        s3_client = get_s3_client(cfg)
        log_and_status(status_fn, f"S3 client initialized (bucket: {s3_bucket})")
    except Exception as e:
        log_error(status_fn, f"Failed to create S3 client: {e}")
        return

    # Check for existing output
    if processing_mode == "skip" and os.path.exists(output_file):
        log_warning(status_fn, f"Output file exists and mode is 'skip': {output_file}")
        log_and_status(status_fn, "Set processing mode to 'overwrite' to regenerate")
        return

    # Read Excel
    rows = _read_excel(input_file, status_fn)
    if not rows:
        log_error(status_fn, "No data rows found in Excel file")
        return

    # Group by parent_sku
    groups = _group_by_parent_sku(rows)
    log_and_status(
        status_fn,
        f"Grouped {len(rows)} rows into {len(groups)} products",
    )

    # Apply start/end record filters
    group_keys = list(groups.keys())
    start_idx = 0
    end_idx = len(group_keys)

    if start_record:
        try:
            start_idx = max(0, int(start_record) - 1)
        except ValueError:
            pass

    if end_record:
        try:
            end_idx = min(len(group_keys), int(end_record))
        except ValueError:
            pass

    group_keys = group_keys[start_idx:end_idx]
    log_and_status(
        status_fn,
        f"Processing products {start_idx + 1} to {start_idx + len(group_keys)} "
        f"of {len(groups)} total",
    )

    # Build products
    products = []
    for i, parent_sku in enumerate(group_keys, start=1):
        group_rows = groups[parent_sku]
        parent_row = _find_parent_row(group_rows, parent_sku)
        log_progress(
            status_fn,
            i, len(group_keys),
            parent_row.get("title", str(parent_sku)),
            details=f"parent_sku={parent_sku}, {len(group_rows)} variant(s)",
        )

        product = _build_product(parent_sku, group_rows, s3_client, s3_bucket, s3_folder, status_fn)
        products.append(product)

    # Write output
    output_data = {"products": products}
    os.makedirs(os.path.dirname(output_file) or ".", exist_ok=True)
    with open(output_file, "w", encoding="utf-8") as f:
        json.dump(output_data, f, indent=2, ensure_ascii=False)

    # Summary
    total_variants = sum(len(p.get("variants", [])) for p in products)
    total_images = sum(len(p.get("images", [])) for p in products)
    log_summary(status_fn, "Processing Complete", {
        "Products": len(products),
        "Total Variants": total_variants,
        "Total Images": total_images,
        "Output File": output_file,
    })
